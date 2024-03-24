import pandas as pd
# Chemin vers le fichier Excel
file_path= 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(file_path)

# Initialiser un DataFrame pour stocker les ROE de toutes les banques pour chaque année
roe_df = pd.DataFrame()

# Itérer sur les années de 2010 à 2022
for year in range(2010, 2023):
    try:
        bilan_df = pd.read_excel(xls, f'BILAN {year}')
        resultat_df = pd.read_excel(xls, f'E Rslt {year}')

        # Utilisation d'une condition pour gérer les variations dans les noms de colonnes ou les libellés
        col_name = 'En milliers de dinars'  # ou adaptez selon les variations observées
        idx_benefice_net = resultat_df.index[resultat_df[col_name].str.contains("RESULTAT NET DE l'EXERCICE", na=False)].tolist()[0]
        idx_fonds_propres = bilan_df.index[bilan_df[col_name].str.contains('TOTAL CAPITAUX PROPRES', na=False)].tolist()[0]

        benefice_net = pd.to_numeric(resultat_df.iloc[idx_benefice_net, 1:].dropna(), errors='coerce')
        fonds_propres = pd.to_numeric(bilan_df.iloc[idx_fonds_propres, 1:].dropna(), errors='coerce')

        roe_percentage = (benefice_net / fonds_propres) * 100
        roe_df[year] = roe_percentage
    except Exception as e:
        print(f"Erreur pour l'année {year}: {e}")

# Transposer le DataFrame pour avoir les années en lignes et les banques en colonnes
roe_df = roe_df.T

print(roe_df)
roe_df.style.format("{:.2f}")

from sklearn.linear_model import LinearRegression
import pandas as pd
import numpy as np

# Load the provided Excel file
file_path= 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(file_path)

# Initialize an empty DataFrame to collect ROE data
roe_df = pd.DataFrame()

# Populate the DataFrame with ROE calculations (this code is based on your initial setup)
for year in range(2010, 2023):
    try:
        bilan_df = pd.read_excel(xls, f'BILAN {year}')
        resultat_df = pd.read_excel(xls, f'E Rslt {year}')
        # Assuming consistent column naming for simplicity
        col_name = 'En milliers de dinars'
        idx_benefice_net = resultat_df.index[resultat_df[col_name].str.contains("RESULTAT NET DE l'EXERCICE", na=False)].tolist()[0]
        idx_fonds_propres = bilan_df.index[bilan_df[col_name].str.contains('TOTAL CAPITAUX PROPRES', na=False)].tolist()[0]
        benefice_net = pd.to_numeric(resultat_df.iloc[idx_benefice_net, 1:].dropna(), errors='coerce')
        fonds_propres = pd.to_numeric(bilan_df.iloc[idx_fonds_propres, 1:].dropna(), errors='coerce')
        roe_percentage = (benefice_net / fonds_propres) * 100
        roe_df[year] = roe_percentage
    except Exception as e:
        print(f"Erreur pour l'année {year}: {e}")

# Transpose the DataFrame for modeling
roe_df = roe_df.T

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression

# Créer un index de prévision avec uniquement les années
forecast_years = range(2023, 2036)  # Crée une liste d'années de 2023 à 2036
forecast_index = pd.Index(forecast_years)

# Définir X_future en utilisant forecast_index pour la prédiction
X_future = np.array(forecast_index).reshape(-1, 1)

# Initialiser un DataFrame pour stocker les prédictions de ROE pour chaque banque de 2023 à 2028
roe_predictions_df = pd.DataFrame(index=forecast_index)

# Appliquer la régression linéaire pour chaque colonne (banque) dans le DataFrame ROE d'origine
for column in roe_df.columns:
    # Extraire les données pour la banque actuelle
    bank_data = roe_df[column].dropna()
    X = np.array(bank_data.index).reshape(-1, 1)  # Années
    y = bank_data.values  # ROE

    # Vérifier s'il existe des valeurs non numériques dans y
    if np.isnan(y).any():
        print(f"Skipping bank '{column}' due to non-numeric values in ROE data.")
        continue

    # Si la banque a suffisamment de données, entraîner le modèle de régression linéaire et prédire
    if len(X) > 1:  # Vérifier s'il y a plus d'une année de données
        lr_model = LinearRegression().fit(X, y)
        predictions = lr_model.predict(X_future)
        predictions = np.round(predictions, 2)  # Arrondir les prédictions à deux chiffres après la virgule

    else:
        # S'il n'y a pas suffisamment de données, remplir avec NaN
        predictions = np.full(len(forecast_years), np.nan)

    # Add predictions to the DataFrame
    roe_predictions_df[column] = predictions

# Display the ROE predictions for each bank from 2023 to 2028
roe_predictions_df
# Transposer le DataFrame pour avoir les années en lignes et les banques en colonnes
output_file_path = 'roepred.xlsx'
roe_predictions_df.to_excel(output_file_path)

# Afficher un aperçu des données de ROA
roe_predictions_df

# Fusionner les DataFrames roe_df et roe_predictions_df
merged_df = pd.concat([roe_df, roe_predictions_df])

# Afficher un aperçu du DataFrame fusionné
print(merged_df.head())  # Afficher seulement les premières lignes du DataFrame fusionné pour un aperçu

# Chemin vers le fichier Excel de sortie pour le DataFrame fusionné
output_merged_file_path = 'merged_roe.xlsx'

# Exporter le DataFrame fusionné au format Excel
merged_df.to_excel(output_merged_file_path)

# Afficher un message de confirmation
print(f"Le DataFrame fusionné a été exporté avec succès au chemin : {output_merged_file_path}")

import matplotlib.pyplot as plt

def evaluate_roe(dataframe):
    # Calcul de la moyenne annuelle du ROE pour chaque banque
    roe_mean = dataframe.mean(axis=0)
    
    # Tracer un graphique pour visualiser l'évolution du ROE moyen au fil des années
    plt.figure(figsize=(10, 6))
    plt.plot(roe_mean.index, roe_mean.values, marker='o', linestyle='-')
    plt.title('Évolution du ROE moyen')
    plt.xlabel('Année')
    plt.ylabel('ROE moyen (%)')
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # Afficher les statistiques descriptives du ROE moyen
    print("\nStatistiques descriptives du ROE moyen :")
    print(roe_mean.describe())

evaluate_roe(merged_df)

from sklearn.linear_model import LinearRegression
import numpy as np
import pandas as pd

# Initialize a dictionary to store models for each bank
models = {}

# Train a model for each bank
for bank in roe_df.columns:
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Years
    y = merged_df[bank].values  # ROE values for the bank
    
    # Linear Regression Model
    model = LinearRegression()
    model.fit(X, y)
    
    # Store the model
    models[bank] = model

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'roe.xlsx'

# Charger les données à partir du fichier Excel
roe_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données
for bank, model in models.items():
    X_real = np.array(range(2010, 2023)).reshape(-1, 1)  # Années avec données réelles
    X_future = np.array(range(2023, 2036)).reshape(-1, 1)  # Années futures sans données réelles
    y_real = roe_df_eval.loc[2010:2022, bank].values  # Valeurs de ROE pour les années avec données réelles
    
    # Prédiction des valeurs de ROE pour les années avec données réelles
    y_pred_real = model.predict(X_real)
    
    # Calcul des métriques d'évaluation uniquement pour les années avec données réelles
    r2 = r2_score(y_real, y_pred_real)
    mse = mean_squared_error(y_real, y_pred_real)
    mae = mean_absolute_error(y_real, y_pred_real)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

    # Prédiction des valeurs de ROE pour les années futures
    y_pred_future = model.predict(X_future)

    # Tracer les valeurs réelles et prédictions (incluant les années futures)
    plt.figure(figsize=(10, 6))
    plt.plot(range(2010, 2023), y_real, label='Valeurs Réelles jusqu\'en 2022')
    plt.plot(range(2023, 2036), y_pred_future, '--', label='Prédictions 2023-2035')
    plt.title(f'Prédictions vs Valeurs Réelles pour {bank}')
    plt.xlabel('Année')
    plt.ylabel('ROE')
    plt.legend()
    plt.show()

# Convertir les évaluations en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)


import pandas as pd
import numpy as np
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'roe.xlsx'

# Charger les données à partir du fichier Excel
roe_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roe_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations)

# Imprimer une explication sur la méthode de calcul de R2
print("Le coefficient de détermination R2 est calculé comme le carré du coefficient de corrélation Pearson entre les valeurs réelles et prédites. Cela mesure la proportion de la variance dans les valeurs réelles qui peut être expliquée par les valeurs prédites.")

print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

import pandas as pd
import numpy as np
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import matplotlib.pyplot as plt

# Chemin vers le fichier Excel à évaluer
output_file_path = 'roe.xlsx'

# Charger les données à partir du fichier Excel
roe_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données pour chaque banque
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roe_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations pour chaque banque
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations individuelles en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T

# Calculer les métriques d'évaluation moyennes pour le DataFrame complet
mean_r2 = evaluations_df['R2'].mean()
mean_mse = evaluations_df['MSE'].mean()
mean_mae = evaluations_df['MAE'].mean()
mean_rmse = evaluations_df['RMSE'].mean()

# Ajouter l'évaluation globale
evaluations['Global'] = {'R2': mean_r2, 'MSE': mean_mse, 'MAE': mean_mae, 'RMSE': mean_rmse}

# Afficher les évaluations
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

print("\nÉvaluation globale moyenne pour toutes les banques :")
print(f"R2 Moyen: {mean_r2}, MSE Moyen: {mean_mse}, MAE Moyen: {mean_mae}, RMSE Moyen: {mean_rmse}")

# Calcul de la corrélation entre les variables
correlation_matrix = roe_df.corr()
print("Matrice de corrélation :")
print(correlation_matrix)

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from sklearn.linear_model import LinearRegression

# Suppose roe_df represents your DataFrame of ROE values for each bank over the years
# This is a simulated structure based on your description. Replace it with your actual DataFrame.
banks = ['ATB', 'BNA', 'ATTIBK', 'ABC', 'BT', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB','BH','BTK','QNB','BTE','BZ','BTS','Secteur']
years = list(range(2010, 2023))
data = np.random.rand(len(years), len(banks)) * 20 + 5  # Simulated ROE values
roe_df = pd.DataFrame(data, columns=banks, index=years)


# Calculate and display correlation matrix
correlation_matrix = roe_df.corr()
print("Correlation matrix between banks' ROEs:")
print(correlation_matrix)

# Plotting the ROE trends

# Function to compute residuals and perform regression for each bank
def compute_and_plot_residuals(dataframe):
    residuals_dict = {}
    for bank in dataframe.columns:
        X = dataframe.index.values.reshape(-1, 1)  # Years
        y = dataframe[bank].values  # ROE values
        
        # Linear regression
        model = LinearRegression().fit(X, y)
        y_pred = model.predict(X)
        
        # Compute residuals
        residuals = y - y_pred
        residuals_dict[bank] = residuals
        
        # Plotting residuals
        plt.figure(figsize=(10, 6))
        plt.scatter(X, residuals, label=f'Residuals for {bank}')
        plt.axhline(y=0, color='red', linestyle='--')
        plt.title(f'Residuals for {bank}')
        plt.xlabel('Year')
        plt.ylabel('Residuals')
        plt.legend()
        plt.show()
        
    return residuals_dict

# Compute and plot residuals for each bank
residuals_dict = compute_and_plot_residuals(roe_df)


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'roe.xlsx'

# Charger les données à partir du fichier Excel
roe_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données pour chaque banque
for bank, model in models.items():
    X_real = np.array(range(2010, 2023)).reshape(-1, 1)  # Années avec données réelles
    X_future = np.array(range(2023, 2036)).reshape(-1, 1)  # Années futures sans données réelles
    y_real = roe_df_eval.loc[2010:2022, bank].values  # Valeurs de ROE pour les années avec données réelles
    
    # Prédiction des valeurs de ROE pour les années avec données réelles
    y_pred_real = model.predict(X_real)
    
    # Calcul des métriques d'évaluation uniquement pour les années avec données réelles
    r2 = r2_score(y_real, y_pred_real)
    mse = mean_squared_error(y_real, y_pred_real)
    mae = mean_absolute_error(y_real, y_pred_real)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

    # Prédiction des valeurs de ROE pour les années futures
    y_pred_future = model.predict(X_future)

    # Tracer les valeurs réelles et prédictions (incluant les années futures)
    plt.figure(figsize=(10, 6))
    plt.plot(range(2010, 2023), y_real, label='Valeurs Réelles jusqu\'en 2022')
    plt.plot(range(2023, 2036), y_pred_future, '--', label='Prédictions 2023-2035')
    plt.title(f'Prédictions vs Valeurs Réelles pour {bank}')
    plt.xlabel('Année')
    plt.ylabel('ROE')
    plt.legend()
    plt.show()

# Convertir les évaluations en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

import matplotlib.pyplot as plt
import numpy as np

# Récupération des méthodes et des scores
methods = list(evaluations_df.index)
r2_scores = evaluations_df['R2']
mse_scores = evaluations_df['MSE']
mae_scores = evaluations_df['MAE']

# Création des positions des barres
bar_width = 0.2
index = np.arange(len(methods))
opacity = 0.8

# Création de la figure et des sous-graphiques
fig, ax = plt.subplots(figsize=(12, 8))

# Barres pour R²
rects1 = ax.bar(index, r2_scores, bar_width, alpha=opacity, color='b', label='R²')

# Barres pour MSE
rects2 = ax.bar(index + bar_width, mse_scores, bar_width, alpha=opacity, color='g', label='MSE')

# Barres pour MAE
rects3 = ax.bar(index + 2 * bar_width, mae_scores, bar_width, alpha=opacity, color='r', label='MAE')

# Ajout des étiquettes, titres et légendes
ax.set_xlabel('Méthodes')
ax.set_ylabel('Scores')
ax.set_title('Comparaison des performances des méthodes de prédiction')
ax.set_xticks(index + bar_width)
ax.set_xticklabels(methods)
ax.legend()

# Affichage du graphique
plt.tight_layout()
plt.show()

import pandas as pd

# Chemin vers le fichier Excel
file_path= 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(file_path)

# Initialiser un DataFrame pour stocker les ROA de toutes les banques pour chaque année
roa_df = pd.DataFrame()

# Itérer sur les années de 2010 à 2022
for year in range(2010, 2023):
    try:
        bilan_df = pd.read_excel(xls, f'BILAN {year}')
        resultat_df = pd.read_excel(xls, f'E Rslt {year}')

        # Utilisation d'une condition pour gérer les variations dans les noms de colonnes ou les libellés
        col_name = 'En milliers de dinars'  # ou adaptez selon les variations observées
        idx_benefice_net = resultat_df.index[resultat_df[col_name].str.contains("RESULTAT NET DE l'EXERCICE", na=False)].tolist()[0]
        idx_total_actifs = bilan_df.index[bilan_df[col_name].str.contains('Total actif', na=False)].tolist()[0]

        benefice_net = pd.to_numeric(resultat_df.iloc[idx_benefice_net, 1:].dropna(), errors='coerce')
        total_actifs = pd.to_numeric(bilan_df.iloc[idx_total_actifs, 1:].dropna(), errors='coerce')

        roa_percentage = (benefice_net / total_actifs) * 100
        roa_df[year] = roa_percentage
    except Exception as e:
        print(f"Erreur pour l'année {year}: {e}")

# Transposer le DataFrame pour avoir les années en lignes et les banques en colonnes
roa_df = roa_df.T

print(roa_df)
roa_df.style.format("{:.2f}")

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression

# Créer un index de prévision avec uniquement les années
forecast_years = range(2023, 2036)  # Crée une liste d'années de 2023 à 2036
forecast_index = pd.Index(forecast_years)

# Définir X_future en utilisant forecast_index pour la prédiction
X_future = np.array(forecast_index).reshape(-1, 1)

# Initialiser un DataFrame pour stocker les prédictions de ROE pour chaque banque de 2023 à 2028
roa_predictions_df = pd.DataFrame(index=forecast_index)

# Appliquer la régression linéaire pour chaque colonne (banque) dans le DataFrame ROE d'origine
for column in roa_df.columns:
    # Extraire les données pour la banque actuelle
    bank_data = roa_df[column].dropna()
    X = np.array(bank_data.index).reshape(-1, 1)  # Années
    y = bank_data.values  # ROE

    # Vérifier s'il existe des valeurs non numériques dans y
    if np.isnan(y).any():
        print(f"Skipping bank '{column}' due to non-numeric values in ROE data.")
        continue

    # Si la banque a suffisamment de données, entraîner le modèle de régression linéaire et prédire
    if len(X) > 1:  # Vérifier s'il y a plus d'une année de données
        lr_model = LinearRegression().fit(X, y)
        predictions = lr_model.predict(X_future)
        predictions = np.round(predictions, 2)  # Arrondir les prédictions à deux chiffres après la virgule

    else:
        # S'il n'y a pas suffisamment de données, remplir avec NaN
        predictions = np.full(len(forecast_years), np.nan)

    # Add predictions to the DataFrame
    roa_predictions_df[column] = predictions

# Display the ROE predictions for each bank from 2023 to 2028
roa_predictions_df
# Transposer le DataFrame pour avoir les années en lignes et les banques en colonnes
output_file_path = 'roapred.xlsx'
roa_predictions_df.to_excel(output_file_path)

# Afficher un aperçu des données de ROA
roa_predictions_df

# Fusionner les DataFrames roe_df et roe_predictions_df
merged_roa_df = pd.concat([roa_df, roa_predictions_df])

# Afficher un aperçu du DataFrame fusionné
print(merged_roa_df.head())  # Afficher seulement les premières lignes du DataFrame fusionné pour un aperçu

# Chemin vers le fichier Excel de sortie pour le DataFrame fusionné
output_merged_file_path = 'merged_roa.xlsx'

# Exporter le DataFrame fusionné au format Excel
merged_roa_df.to_excel(output_merged_file_path)

# Afficher un message de confirmation
print(f"Le DataFrame fusionné a été exporté avec succès au chemin : {output_merged_file_path}")

import matplotlib.pyplot as plt

def evaluate_roa(dataframe):
    # Calcul de la moyenne annuelle du ROE pour chaque banque
    roa_mean = dataframe.mean(axis=0)
    
    # Tracer un graphique pour visualiser l'évolution du ROE moyen au fil des années
    plt.figure(figsize=(10, 6))
    plt.plot(roa_mean.index, roa_mean.values, marker='o', linestyle='-')
    plt.title('Évolution du ROE moyen')
    plt.xlabel('Année')
    plt.ylabel('ROE moyen (%)')
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # Afficher les statistiques descriptives du ROE moyen
    print("\nStatistiques descriptives du ROE moyen :")
    print(roa_mean.describe())

evaluate_roa(merged_roa_df)

from sklearn.linear_model import LinearRegression
import numpy as np
import pandas as pd

# Initialize a dictionary to store models for each bank
models = {}

# Train a model for each bank
for bank in roa_df.columns:
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Years
    y = merged_roa_df[bank].values  # ROE values for the bank
    
    # Linear Regression Model
    model = LinearRegression()
    model.fit(X, y)
    
    # Store the model
    models[bank] = model

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'merged_roa.xlsx'

# Charger les données à partir du fichier Excel
roa_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données
for bank, model in models.items():
    X_real = np.array(range(2010, 2023)).reshape(-1, 1)  # Années avec données réelles
    X_future = np.array(range(2023, 2036)).reshape(-1, 1)  # Années futures sans données réelles
    y_real = roa_df_eval.loc[2010:2022, bank].values  # Valeurs de ROE pour les années avec données réelles
    
    # Prédiction des valeurs de ROE pour les années avec données réelles
    y_pred_real = model.predict(X_real)
    
    # Calcul des métriques d'évaluation uniquement pour les années avec données réelles
    r2 = r2_score(y_real, y_pred_real)
    mse = mean_squared_error(y_real, y_pred_real)
    mae = mean_absolute_error(y_real, y_pred_real)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

    # Prédiction des valeurs de ROE pour les années futures
    y_pred_future = model.predict(X_future)

    # Tracer les valeurs réelles et prédictions (incluant les années futures)
    plt.figure(figsize=(10, 6))
    plt.plot(range(2010, 2023), y_real, label='Valeurs Réelles jusqu\'en 2022')
    plt.plot(range(2023, 2036), y_pred_future, '--', label='Prédictions 2023-2035')
    plt.title(f'Prédictions vs Valeurs Réelles pour {bank}')
    plt.xlabel('Année')
    plt.ylabel('ROE')
    plt.legend()
    plt.show()

# Convertir les évaluations en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)


import pandas as pd
import numpy as np
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'merged_roa.xlsx'

# Charger les données à partir du fichier Excel
roa_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roa_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations)

# Imprimer une explication sur la méthode de calcul de R2
print("Le coefficient de détermination R2 est calculé comme le carré du coefficient de corrélation Pearson entre les valeurs réelles et prédites. Cela mesure la proportion de la variance dans les valeurs réelles qui peut être expliquée par les valeurs prédites.")

print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

import pandas as pd
import numpy as np
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import matplotlib.pyplot as plt

# Chemin vers le fichier Excel à évaluer
output_file_path =  'merged_roa.xlsx'

# Charger les données à partir du fichier Excel
roa_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données pour chaque banque
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roa_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations pour chaque banque
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations individuelles en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T

# Calculer les métriques d'évaluation moyennes pour le DataFrame complet
mean_r2 = evaluations_df['R2'].mean()
mean_mse = evaluations_df['MSE'].mean()
mean_mae = evaluations_df['MAE'].mean()
mean_rmse = evaluations_df['RMSE'].mean()

# Ajouter l'évaluation globale
evaluations['Global'] = {'R2': mean_r2, 'MSE': mean_mse, 'MAE': mean_mae, 'RMSE': mean_rmse}

# Afficher les évaluations
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

print("\nÉvaluation globale moyenne pour toutes les banques :")
print(f"R2 Moyen: {mean_r2}, MSE Moyen: {mean_mse}, MAE Moyen: {mean_mae}, RMSE Moyen: {mean_rmse}")

# Calcul de la corrélation entre les variables
correlation_matrix = roa_df.corr()
print("Matrice de corrélation :")
print(correlation_matrix)

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from sklearn.linear_model import LinearRegression

# Suppose roe_df represents your DataFrame of ROE values for each bank over the years
# This is a simulated structure based on your description. Replace it with your actual DataFrame.
banks = ['ATB', 'BNA', 'ATTIBK', 'ABC', 'BT', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB','BH','BTK','QNB','BTE','BZ','BTS','Secteur']
years = list(range(2010, 2023))
data = np.random.rand(len(years), len(banks)) * 20 + 5  # Simulated ROE values
roa_df = pd.DataFrame(data, columns=banks, index=years)


# Calculate and display correlation matrix
correlation_matrix = roa_df.corr()
print("Correlation matrix between banks' ROEs:")
print(correlation_matrix)

# Plotting the ROE trends

# Function to compute residuals and perform regression for each bank
def compute_and_plot_residuals(dataframe):
    residuals_dict = {}
    for bank in dataframe.columns:
        X = dataframe.index.values.reshape(-1, 1)  # Years
        y = dataframe[bank].values  # ROE values
        
        # Linear regression
        model = LinearRegression().fit(X, y)
        y_pred = model.predict(X)
        
        # Compute residuals
        residuals = y - y_pred
        residuals_dict[bank] = residuals
        
        # Plotting residuals
        plt.figure(figsize=(10, 6))
        plt.scatter(X, residuals, label=f'Residuals for {bank}')
        plt.axhline(y=0, color='red', linestyle='--')
        plt.title(f'Residuals for {bank}')
        plt.xlabel('Year')
        plt.ylabel('Residuals')
        plt.legend()
        plt.show()
        
    return residuals_dict

# Compute and plot residuals for each bank
residuals_dict = compute_and_plot_residuals(roa_df)


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'merged_roa.xlsx'

# Charger les données à partir du fichier Excel
roa_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données pour chaque banque
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roa_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations pour chaque banque
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations individuelles en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T

# Calculer les métriques d'évaluation moyennes pour le DataFrame complet
mean_r2 = evaluations_df['R2'].mean()
mean_mse = evaluations_df['MSE'].mean()
mean_mae = evaluations_df['MAE'].mean()
mean_rmse = evaluations_df['RMSE'].mean()

# Ajouter l'évaluation globale
evaluations['Global'] = {'R2': mean_r2, 'MSE': mean_mse, 'MAE': mean_mae, 'RMSE': mean_rmse}

# Afficher les évaluations
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

print("\nÉvaluation globale moyenne pour toutes les banques :")
print(f"R2 Moyen: {mean_r2}, MSE Moyen: {mean_mse}, MAE Moyen: {mean_mae}, RMSE Moyen: {mean_rmse}")

# Tracer les métriques d'évaluation
plt.figure(figsize=(10, 6))
evaluations_df.plot(kind='bar', figsize=(10, 6))
plt.title('Évaluation du ROE pour chaque banque')
plt.xlabel('Banque')
plt.ylabel('Métrique')
plt.xticks(rotation=45)
plt.legend(title='Métrique')
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
plt.show()

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

# Chemin vers le fichier Excel à évaluer
output_file_path = 'merged_roa.xlsx'

# Charger les données à partir du fichier Excel
roa_df_eval = pd.read_excel(output_file_path, index_col=0)

# Initialiser un dictionnaire pour stocker les évaluations pour chaque banque
evaluations = {}

# Évaluer les données pour chaque banque
for bank, model in models.items():
    X = np.array(range(2010, 2036)).reshape(-1, 1)  # Années
    y = roa_df_eval[bank].values  # Valeurs de ROE pour la banque
    
    # Prédiction des valeurs de ROE
    y_pred = model.predict(X)
    
    # Calcul des métriques d'évaluation
    r2 = r2_score(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    rmse = np.sqrt(mse)
    
    # Stocker les évaluations pour chaque banque
    evaluations[bank] = {'R2': r2, 'MSE': mse, 'MAE': mae, 'RMSE': rmse}

# Convertir les évaluations individuelles en DataFrame pour affichage
evaluations_df = pd.DataFrame(evaluations).T

# Calculer les métriques d'évaluation moyennes pour le DataFrame complet
mean_r2 = evaluations_df['R2'].mean()
mean_mse = evaluations_df['MSE'].mean()
mean_mae = evaluations_df['MAE'].mean()
mean_rmse = evaluations_df['RMSE'].mean()

# Ajouter l'évaluation globale
evaluations['Global'] = {'R2': mean_r2, 'MSE': mean_mse, 'MAE': mean_mae, 'RMSE': mean_rmse}

# Afficher les évaluations
print("\nÉvaluation du ROE des données actuelles pour chaque banque :")
print(evaluations_df)

print("\nÉvaluation globale moyenne pour toutes les banques :")
print(f"R2 Moyen: {mean_r2}, MSE Moyen: {mean_mse}, MAE Moyen: {mean_mae}, RMSE Moyen: {mean_rmse}")

# Tracer les métriques d'évaluation
plt.figure(figsize=(10, 6))
evaluations_df.plot(kind='bar', figsize=(10, 6))
plt.title('Évaluation du ROE pour chaque banque')
plt.xlabel('Banque')
plt.ylabel('Métrique')
plt.xticks(rotation=45)
plt.legend(title='Métrique')
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
plt.show()
import pandas as pd
import matplotlib.pyplot as plt

# Chargement du fichier Excel
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)

# Initialisation d'un DataFrame pour stocker les résultats
results_df = pd.DataFrame()

for year in range(2010, 2023):  # Pour chaque année de 2010 à 2022
    bilan_sheet_name = f'BILAN {year}'
    resultat_sheet_name = f'E Rslt {year}'
    
    if bilan_sheet_name in xl.sheet_names and resultat_sheet_name in xl.sheet_names:
        # Chargement des données des feuilles de bilan et de résultat pour l'année courante
        bilan_data = pd.read_excel(file_path, sheet_name=bilan_sheet_name, usecols=["Secteur"])
        resultat_data = pd.read_excel(file_path, sheet_name=resultat_sheet_name, usecols=["Secteur"])
        
        # Calcul des ratios en utilisant les premières lignes pour une approximation
        actifs_courants_approx = bilan_data.iloc[0]['Secteur']
        passifs_courants_approx = bilan_data.iloc[1]['Secteur']
        capitaux_propres_approx = bilan_data.iloc[2]['Secteur']  # Approximation pour les capitaux propres
        resultat_exploitation_approx = resultat_data.iloc[0]['Secteur']
        charges_interet_approx = resultat_data.iloc[1]['Secteur']
        
        ratio_liquidite_courante_approx = actifs_courants_approx / passifs_courants_approx if passifs_courants_approx else pd.NA
        ratio_dette_sur_fonds_propres_approx = (actifs_courants_approx + passifs_courants_approx) / capitaux_propres_approx if capitaux_propres_approx else pd.NA
        ratio_couverture_interets_approx = resultat_exploitation_approx / charges_interet_approx if charges_interet_approx else pd.NA
        
        # Ajout des résultats approximatifs au DataFrame
        results_df = pd.concat([results_df, pd.DataFrame({
            'Année': [year],
            'Ratio de Liquidité Courante Approximatif': [ratio_liquidite_courante_approx],
            'Ratio Dette sur Fonds Propres Approximatif': [ratio_dette_sur_fonds_propres_approx],
            'Ratio de Couverture des Intérêts Approximatif': [ratio_couverture_interets_approx]
        })], ignore_index=True)

# Configuration de 'Année' comme index du DataFrame
results_df.set_index('Année', inplace=True)

# Affichage du DataFrame avec les résultats
print(results_df)

# Génération des graphiques pour visualiser les résultats
fig, axs = plt.subplots(3, 1, figsize=(10, 15))

results_df['Ratio de Liquidité Courante Approximatif'].plot(kind='bar', ax=axs[0], color='skyblue')
axs[0].set_title('Ratio de Liquidité Courante Approximatif par Année')
axs[0].set_ylabel('Ratio')

results_df['Ratio Dette sur Fonds Propres Approximatif'].plot(kind='bar', ax=axs[1], color='lightgreen')
axs[1].set_title('Ratio Dette sur Fonds Propres Approximatif par Année')
axs[1].set_ylabel('Ratio')

results_df['Ratio de Couverture des Intérêts Approximatif'].plot(kind='bar', ax=axs[2], color='salmon')
axs[2].set_title('Ratio de Couverture des Intérêts Approximatif par Année')
axs[2].set_ylabel('Ratio')

plt.tight_layout()
plt.show()

import pandas as pd
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns

# Chargement du fichier Excel
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)

# Initialisation d'un DataFrame pour stocker les résultats
results_df = pd.DataFrame()

for year in range(2010, 2023):  # Pour chaque année de 2010 à 2022
    bilan_sheet_name = f'BILAN {year}'
    resultat_sheet_name = f'E Rslt {year}'
    
    if bilan_sheet_name in xl.sheet_names and resultat_sheet_name in xl.sheet_names:
        # Chargement des données des feuilles de bilan et de résultat pour l'année courante
        bilan_data = pd.read_excel(file_path, sheet_name=bilan_sheet_name, usecols=["Secteur"])
        resultat_data = pd.read_excel(file_path, sheet_name=resultat_sheet_name, usecols=["Secteur"])
        
        # Calcul des ratios en utilisant les premières lignes pour une approximation
        actifs_courants_approx = bilan_data.iloc[0]['Secteur']
        passifs_courants_approx = bilan_data.iloc[1]['Secteur']
        capitaux_propres_approx = bilan_data.iloc[2]['Secteur']
        resultat_exploitation_approx = resultat_data.iloc[0]['Secteur']
        charges_interet_approx = resultat_data.iloc[1]['Secteur']
        
        ratio_liquidite_courante_approx = actifs_courants_approx / passifs_courants_approx if passifs_courants_approx else pd.NA
        ratio_dette_sur_fonds_propres_approx = (actifs_courants_approx + passifs_courants_approx) / capitaux_propres_approx if capitaux_propres_approx else pd.NA
        ratio_couverture_interets_approx = resultat_exploitation_approx / charges_interet_approx if charges_interet_approx else pd.NA
        
        # Ajout des résultats approximatifs au DataFrame
        results_df = pd.concat([results_df, pd.DataFrame({
            'Année': [year],
            'Ratio de Liquidité Courante Approximatif': [ratio_liquidite_courante_approx],
            'Ratio Dette sur Fonds Propres Approximatif': [ratio_dette_sur_fonds_propres_approx],
            'Ratio de Couverture des Intérêts Approximatif': [ratio_couverture_interets_approx]
        })], ignore_index=True)

# Normalisation des données avant le clustering
scaler = StandardScaler()
scaled_features = scaler.fit_transform(results_df[['Ratio de Liquidité Courante Approximatif', 'Ratio Dette sur Fonds Propres Approximatif', 'Ratio de Couverture des Intérêts Approximatif']].fillna(0))

# Application de K-Means
kmeans = KMeans(n_clusters=3, random_state=42)
kmeans.fit(scaled_features)
results_df['Cluster'] = kmeans.labels_

# Regroupement des années par cluster
results_df['Année'] = results_df['Année'].astype(str)  # Assurer que l'année est en format string pour l'affichage
grouped_years = results_df.groupby('Cluster')['Année'].apply(lambda x: ', '.join(x)).reset_index()

# Affichage des années regroupées par cluster
for index, row in grouped_years.iterrows():
    print(f"Cluster {row['Cluster']} contient les années suivantes : {row['Année']}")

# Optionnel: Visualisation des clusters
sns.pairplot(results_df, hue='Cluster', vars=['Ratio de Liquidité Courante Approximatif', 'Ratio Dette sur Fonds Propres Approximatif', 'Ratio de Couverture des Intérêts Approximatif'])
plt.title('Visualisation des Clusters')
plt.show()

from tabulate import tabulate

# Création de DataFrames pour chaque cluster
cluster_tables = []

for cluster_id in range(kmeans.n_clusters):
    cluster_years = grouped_years.iloc[cluster_id]['Année'].split(', ')
    cluster_values = results_df[results_df['Cluster'] == cluster_id].drop(columns=['Cluster'])
    cluster_data = []
    for year in cluster_years:
        row = cluster_values[cluster_values['Année'] == year].squeeze().to_dict()
        row['Année'] = year
        cluster_data.append(row)

    cluster_table = tabulate(cluster_data, headers='keys', tablefmt='grid')
    cluster_tables.append(cluster_table)

# Ajouter un message pour le cluster manquant
if len(cluster_tables) < kmeans.n_clusters:
    cluster_tables.append("Aucune donnée disponible pour le cluster 2.")

# Affichage des tableaux pour chaque cluster
for i, table in enumerate(cluster_tables):
    print(f"Cluster {i}:")
    print(table)
    print("\n")

from tabulate import tabulate

# Récupération des années associées au cluster 2
cluster_years = grouped_years[grouped_years['Cluster'] == 2]['Année'].values[0].split(', ')

# Récupération des données pour le cluster 2
cluster_values = results_df[results_df['Cluster'] == 2].drop(columns=['Cluster'])

# Création de la table pour le cluster 2
cluster_data = []
for year in cluster_years:
    row = cluster_values[cluster_values['Année'] == year].squeeze().to_dict()
    row['Année'] = year
    cluster_data.append(row)

cluster_table = tabulate(cluster_data, headers='keys', tablefmt='grid')

# Affichage de la table pour le cluster 2
print("Cluster 2:")
print(cluster_table)
import openpyxl

# Charger le fichier Excel
fichier_excel_path = 'EtudeSB - Copie.xlsx'
fichier_excel = openpyxl.load_workbook(fichier_excel_path)

# Initialisation d'un dictionnaire pour stocker les totaux de prêts par année
totals_prets_numeriques_par_annee = {}

# Parcourir les années de 2010 à 2023 (pour inclure 2022)
for annee in range(2010, 2023):
    feuille_bilan = f'BILAN {annee}'
    # Vérifier si la feuille de bilan pour l'année donnée existe dans le fichier
    if feuille_bilan in fichier_excel.sheetnames:
        ws = fichier_excel[feuille_bilan]
        
        # Parcourir les lignes pour trouver "Créances sur clientèle"
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Créances sur clientèle" in row[0]:
                # Trouver la ligne des "Créances sur clientèle"
                # Supposer que le total sectoriel est la dernière valeur numérique de la ligne
                # Calculer le total des prêts pour l'année en sommant les valeurs numériques de la ligne,
                # en ignorant les cellules vides et la formule à la fin
                total_numerique = sum(val for val in row[1:] if isinstance(val, (int, float)))
                # Stocker le total dans le dictionnaire avec l'année comme clé
                totals_prets_numeriques_par_annee[annee] = total_numerique
                break  # Sortir de la boucle une fois la ligne trouvée

# Affichage des résultats
for annee, total in totals_prets_numeriques_par_annee.items():
    print(f"Total des prêts pour {annee}: {total} milliers de dinars")


import pandas as pd
import matplotlib.pyplot as plt

# Créer un DataFrame à partir des totaux des prêts
df_prets = pd.DataFrame(list(totals_prets_numeriques_par_annee.items()), columns=['Année', 'Total des Prêts (en milliers de dinars)'])

# Affichage du DataFrame
print(df_prets)

# Générer un graphique linéaire pour visualiser l'évolution des prêts de 2010 à 2022
plt.figure(figsize=(10, 6))
plt.plot(df_prets['Année'], df_prets['Total des Prêts (en milliers de dinars)'], marker='o', linestyle='-', color='blue')
plt.title('Évolution des prêts dans le secteur bancaire en Tunisie (2010-2022)')
plt.xlabel('Année')
plt.ylabel('Total des Prêts (en milliers de dinars)')
plt.grid(True)
plt.xticks(rotation=45)
plt.tight_layout()

# Sauvegarder le graphique
#graphique_path = " evolution_prets_2010_2022.png"
#plt.savefig(graphique_path)

plt.show() #graphique_path

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error
import numpy as np

# Charger le fichier Excel (Ajustez le chemin selon votre environnement)
fichier_excel_path = 'EtudeSB - Copie.xlsx'
fichier_excel = openpyxl.load_workbook(fichier_excel_path)

totals_prets_numeriques_par_annee = {}

for annee in range(2010, 2023):
    feuille_bilan = f'BILAN {annee}'
    if feuille_bilan in fichier_excel.sheetnames:
        ws = fichier_excel[feuille_bilan]
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Créances sur clientèle" in row[0]:
                total_numerique = sum(val for val in row[1:] if isinstance(val, (int, float)))
                totals_prets_numeriques_par_annee[annee] = total_numerique
                break

# Création d'un DataFrame
df_prets = pd.DataFrame(list(totals_prets_numeriques_par_annee.items()), columns=['Année', 'Total des Prêts'])

# Modélisation: Régression Linéaire
X = df_prets[['Année']]
y = df_prets['Total des Prêts']

# Diviser les données en ensemble d'apprentissage et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

# Entraîner le modèle
modele = LinearRegression()
modele.fit(X_train, y_train)

# Prédiction sur l'ensemble de test
predictions = modele.predict(X_test)

# Evaluation du modèle
mae = mean_absolute_error(y_test, predictions)
mse = mean_squared_error(y_test, predictions)
rmse = np.sqrt(mse)

print(f"MAE: {mae}, MSE: {mse}, RMSE: {rmse}")

# Visualisation des prêts réels vs prédits
plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='black', label='Données réelles')
plt.plot(X_test, predictions, color='blue', linewidth=2, label='Prédiction linéaire')
plt.title('Réel vs Prédit')
plt.xlabel('Année')
plt.ylabel('Total des Prêts')
plt.legend()
plt.show()

from sklearn.model_selection import train_test_split

# Préparation des variables indépendantes (X) et dépendantes (y)
X = df_prets[['Année']].values # Variable indépendante
y = df_prets['Total des Prêts'].values # Variable dépendante

# Séparation des données en ensembles d'entraînement et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

from sklearn.linear_model import LinearRegression

# Création et entraînement du modèle
modele = LinearRegression()
modele.fit(X_train, y_train)

from sklearn.metrics import mean_squared_error, r2_score

# Prédiction sur les données de test
y_pred = modele.predict(X_test)

# Calcul et affichage des métriques d'évaluation
rmse = mean_squared_error(y_test, y_pred, squared=False)
r2 = r2_score(y_test, y_pred)

print(f"RMSE: {rmse}")
print(f"R² Score: {r2}")

plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='black', label='Données réelles')
plt.plot(X_test, y_pred, color='blue', linewidth=3, label='Prédiction linéaire')
plt.title('Prédiction vs Réelles')
plt.xlabel('Année')
plt.ylabel('Total des Prêts (en milliers de dinars)')
plt.legend()
plt.show()
import pandas as pd
import matplotlib.pyplot as plt
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)
# Function to extract financial leverage data from the balance sheet
# This function will try to handle cases where the lookup fails by returning None
def extract_leverage_data_updated(sheet_name, file_path):
    data = pd.read_excel(file_path, sheet_name=sheet_name)
    try:
        # Attempting to extract 'Total passifs' and 'TOTAL CAPITAUX PROPRES'
        total_liabilities = data.loc[data['En milliers de dinars'].str.contains('Total passifs', na=False), 'Secteur'].values[0]
        total_equity = data.loc[data['En milliers de dinars'].str.contains('TOTAL CAPITAUX PROPRES', na=False), 'Secteur'].values[0]
        return total_liabilities, total_equity
    except IndexError:
        # If the expected rows are not found, return None to indicate failure
        return None, None

# Initialize a DataFrame to store leverage results
leverage_results_updated = pd.DataFrame(columns=['Year', 'Financial Leverage'])

# Loop through each year and calculate the financial leverage
for year in range(2010, 2023):
    bilan_sheet_name = f'BILAN {year}'
    # Check if the sheet exists in the Excel file
    if bilan_sheet_name in xl.sheet_names:
        total_liabilities, total_equity = extract_leverage_data_updated(bilan_sheet_name, file_path)
        if total_liabilities is not None and total_equity is not None:
            # Calculate financial leverage if data was successfully extracted
            financial_leverage = total_liabilities / total_equity
            # Append to results DataFrame
            leverage_results_updated = pd.concat([leverage_results_updated, pd.DataFrame({
                'Year': [year],
                'Financial Leverage': [financial_leverage]
            })], ignore_index=True)

# Set 'Year' as the index
leverage_results_updated.set_index('Year', inplace=True)


# Définir le chemin complet du fichier Excel de sortie pour les résultats de l'effet de levier financier
output_leverage_excel_path = 'leverage_results.xlsx'

# Exporter le DataFrame des résultats de l'effet de levier financier dans un fichier Excel
leverage_results_updated.to_excel(output_leverage_excel_path)

# Plotting the financial leverage over the years
plt.figure(figsize=(10, 6))
plt.bar(leverage_results_updated.index, leverage_results_updated['Financial Leverage'], color='skyblue')
plt.title('Financial Leverage from 2010 to 2022')
plt.xlabel('Year')
plt.ylabel('Financial Leverage')
plt.grid(True)
plt.xticks(range(2010, 2023))
plt.tight_layout()
plt.show()

# Display the results as a table
print("Tableau des résultats de l'effet de levier financier :")
print(leverage_results_updated)

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt

# Charger les données
file_path = 'leverage_results.xlsx'
leverage_data = pd.read_excel(file_path, index_col='Year')

# Préparation des données
X = leverage_data.index.values.reshape(-1, 1)  # Années comme caractéristique
y = leverage_data['Financial Leverage'].values  # Effet de levier financier comme cible

# Division des données en ensembles d'entraînement et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Créer et entraîner le modèle de régression linéaire
model = LinearRegression()
model.fit(X_train, y_train)

# Prédiction sur l'ensemble de test
y_pred = model.predict(X_test)

# Évaluation du modèle
rmse = mean_squared_error(y_test, y_pred, squared=False)
r2 = r2_score(y_test, y_pred)

# Affichage des résultats
print(f"RMSE sur l'ensemble de test: {rmse}")
print(f"R2 sur l'ensemble de test: {r2}")

# Tracé des résultats
plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='blue', label='Valeurs réelles')
plt.plot(X_test, y_pred, color='red', label='Prédiction linéaire')
plt.title('Prédiction de l\'effet de levier financier')
plt.xlabel('Année')
plt.ylabel('Effet de levier financier')
plt.legend()
plt.grid(True)
plt.show()

from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt

# Créer et entraîner le modèle de forêt aléatoire
model_rf = RandomForestRegressor(n_estimators=100, random_state=42)  # 100 arbres dans la forêt
model_rf.fit(X_train, y_train)

# Prédiction sur l'ensemble de test avec le modèle de forêt aléatoire
y_pred_rf = model_rf.predict(X_test)

# Évaluation du modèle de forêt aléatoire
rmse_rf = mean_squared_error(y_test, y_pred_rf, squared=False)
r2_rf = r2_score(y_test, y_pred_rf)

# Affichage des résultats pour le modèle de forêt aléatoire
print(f"RMSE (Forêt Aléatoire) sur l'ensemble de test: {rmse_rf}")
print(f"R2 (Forêt Aléatoire) sur l'ensemble de test: {r2_rf}")

# Tracé des résultats
plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='blue', label='Valeurs réelles')
plt.scatter(X_test, y_pred_rf, color='red', label='Prédiction Forêt Aléatoire')
plt.title('Prédiction de l\'effet de levier financier avec Forêt Aléatoire')
plt.xlabel('Année')
plt.ylabel('Effet de levier financier')
plt.legend()
plt.grid(True)
plt.show()


from xgboost import XGBRegressor
from sklearn.metrics import mean_squared_error, r2_score

# Création du modèle XGBoost
model_xgb = XGBRegressor(n_estimators=100, learning_rate=0.1, random_state=42)

# Entraînement du modèle
model_xgb.fit(X_train, y_train)

# Prédictions
y_pred_xgb = model_xgb.predict(X_test)

# Évaluation
rmse_xgb = np.sqrt(mean_squared_error(y_test, y_pred_xgb))
r2_xgb = r2_score(y_test, y_pred_xgb)

print(f"RMSE (XGBoost) sur l'ensemble de test: {rmse_xgb}")
print(f"R2 (XGBoost) sur l'ensemble de test: {r2_xgb}")


import pandas as pd
import matplotlib.pyplot as plt
from statsmodels.tsa.seasonal import seasonal_decompose

# Assurez-vous que 'Year' est de type datetime si ce n'est pas déjà le cas
leverage_results_updated.index = pd.to_datetime(leverage_results_updated.index, format='%Y')

# Effectuer la décomposition saisonnière
result = seasonal_decompose(leverage_results_updated['Financial Leverage'], model='additive')

# Tracer les composants décomposés
fig = result.plot()
fig.set_size_inches(10, 8)
plt.show()

from statsmodels.tsa.stattools import adfuller

result = adfuller(leverage_results_updated['Financial Leverage'])
print('ADF Statistic: %f' % result[0])
print('p-value: %f' % result[1])

# Interprétation
if result[1] > 0.05:
    print("La série n'est pas stationnaire.")
else:
    print("La série est stationnaire.")

from statsmodels.graphics.tsaplots import plot_acf, plot_pacf

# Tracé de l'ACF et PACF
fig, axes = plt.subplots(1, 2, figsize=(15, 4))
plot_acf(leverage_results_updated['Financial Leverage'], ax=axes[0])
plot_pacf(leverage_results_updated['Financial Leverage'], ax=axes[1])
plt.show()

from statsmodels.tsa.arima.model import ARIMA
# Ou SARIMAX pour SARIMA
from statsmodels.tsa.statespace.sarimax import SARIMAX

from statsmodels.tsa.arima.model import ARIMA

p = 1
d = 0
q = 1

model = ARIMA(leverage_results_updated['Financial Leverage'], order=(p, d, q))
model_fit = model.fit()

# Affichage du résumé du modèle
print(model_fit.summary())



# Prévisions
forecast = model_fit.forecast(steps=5)  # Prévoir 5 périodes dans le futur
print(forecast)

import numpy as np
from sklearn.metrics import mean_squared_error, r2_score

# Obtenir les prédictions rétrospectives (in-sample forecast)
# Cela nous donnera les prédictions pour l'ensemble des données utilisées pour l'entraînement du modèle
in_sample_forecast = model_fit.predict(start=leverage_results_updated.index.min(), end=leverage_results_updated.index.max())

# Calculer le RMSE sur l'ensemble d'entraînement
rmse = np.sqrt(mean_squared_error(leverage_results_updated['Financial Leverage'], in_sample_forecast))
print("RMSE sur l'ensemble d'entraînement:", rmse)

# Calculer le R² sur l'ensemble d'entraînement
r2 = r2_score(leverage_results_updated['Financial Leverage'], in_sample_forecast)
print("R² sur l'ensemble d'entraînement:", r2)

# Tracer les valeurs réelles contre les prédictions rétrospectives
plt.figure(figsize=(10, 6))
plt.plot(leverage_results_updated.index, leverage_results_updated['Financial Leverage'], label='Valeurs Réelles')
plt.plot(leverage_results_updated.index, in_sample_forecast, label='Prédictions Rétrospectives', linestyle='--')
plt.title('Prédictions Rétrospectives avec le Modèle ARIMA')
plt.xlabel('Année')
plt.ylabel('Effet de levier financier')
plt.legend()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
import os
from pathlib import Path

# Define the path to the Excel file
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)

# Helper function to extract financial data
def extract_financial_data(sheet_name, file_path):
    data = pd.read_excel(file_path, sheet_name=sheet_name)
    metrics = {}
    try:
        metrics['total_liabilities'] = data.loc[data['En milliers de dinars'].str.contains('Total passifs', na=False), 'Secteur'].values[0]
        metrics['total_equity'] = data.loc[data['En milliers de dinars'].str.contains('TOTAL CAPITAUX PROPRES', na=False), 'Secteur'].values[0]
        metrics['total_assets'] = metrics['total_liabilities'] + metrics['total_equity']
        # Add more financial metrics if needed
    except IndexError:
        metrics = None
    return metrics

# Initialize a DataFrame to store enhanced financial analysis results
financial_analysis_results = pd.DataFrame()

# Loop through each sheet in the Excel file
for sheet_name in xl.sheet_names:
    if sheet_name.startswith('BILAN'):
        financial_data = extract_financial_data(sheet_name, file_path)
        if financial_data:
            year = int(sheet_name.split(' ')[1])
            financial_leverage = financial_data['total_liabilities'] / financial_data['total_equity']
            debt_to_equity = financial_data['total_liabilities'] / financial_data['total_equity']
            # Append to results DataFrame
            financial_analysis_results = pd.concat([financial_analysis_results, pd.DataFrame({
                'Year': [year],
                'Financial Leverage': [financial_leverage],
                'Debt-to-Equity Ratio': [debt_to_equity]
            })], ignore_index=True)

# Sort the results DataFrame by 'Year'
financial_analysis_results.sort_values('Year', inplace=True)
financial_analysis_results.set_index('Year', inplace=True)

# Export the enhanced financial analysis to a new Excel file
output_path = 'enhanced_financial_analysis.xlsx' 
financial_analysis_results.to_excel(output_path)

# Plotting multiple financial metrics over the years
plt.figure(figsize=(14, 8))
plt.subplot(311)
plt.plot(financial_analysis_results.index, financial_analysis_results['Financial Leverage'], marker='o', linestyle='-', color='blue')
plt.title('Financial Leverage over Years')
plt.xlabel('Year')
plt.ylabel('Financial Leverage')
plt.grid(True)


plt.subplot(313)
plt.plot(financial_analysis_results.index, financial_analysis_results['Debt-to-Equity Ratio'], marker='o', linestyle='-', color='green')
plt.title('Debt-to-Equity Ratio over Years')
plt.xlabel('Year')
plt.ylabel('Debt-to-Equity Ratio')
plt.grid(True)

plt.tight_layout()
plt.show()

# Return the path to the saved Excel file for download
output_path

from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import make_pipeline

# Simulons des données pour l'exemple, remplacez ceci par votre propre préparation de données
X = financial_analysis_results.drop('Financial Leverage', axis=1)
y = financial_analysis_results['Financial Leverage']

# Division des données en ensembles d'entraînement et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Modèle
model = make_pipeline(StandardScaler(), RandomForestRegressor(n_estimators=100, random_state=42))
model.fit(X_train, y_train)

# Prédictions et évaluation
y_pred = model.predict(X_test)
r2 = r2_score(y_test, y_pred)

print(f"R^2: {r2}")

# Optimisation: Cette étape peut inclure l'ajustement des hyperparamètres, le choix d'un modèle différent, ou la transformation des caractéristiques.

import numpy as np
from sklearn.linear_model import LinearRegression
import pandas as pd
import matplotlib.pyplot as plt
# Assuming `financial_analysis_results` is your DataFrame with historical data
# For demonstration, let's create a simple DataFrame
years = np.arange(2010, 2023).reshape(-1, 1)  # Example years
financial_leverage = np.random.rand(len(years))  # Example financial leverage data

# Fit the model
model = LinearRegression().fit(years, financial_leverage)
future_years = np.arange(2023, 2036).reshape(-1, 1)  # Years to predict
predicted_leverage = model.predict(future_years)

# Create a DataFrame for the predictions
predictions_df = pd.DataFrame({
    'Year': future_years.flatten(),
    'Predicted Financial Leverage': predicted_leverage
})
# Plot historical data
plt.plot(years, financial_leverage, label='Historical Financial Leverage')

# Plot predicted data
plt.plot(future_years, predicted_leverage, label='Predicted Financial Leverage', linestyle='--')

plt.xlabel('Year')
plt.ylabel('Financial Leverage')
plt.title('Financial Leverage Prediction from 2023 to 2035')
plt.legend()
plt.show()

import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error

# Example historical data (simulated)
np.random.seed(0)  # For reproducible output
years = np.arange(2000, 2023).reshape(-1, 1)
financial_leverage = np.random.rand(len(years)) * 0.1 + np.linspace(0.5, 1.5, len(years))

# Fit the linear regression model
model = LinearRegression()
model.fit(years, financial_leverage)
# Years to predict
future_years = np.arange(2023, 2036).reshape(-1, 1)

# Make predictions
predicted_leverage = model.predict(future_years)

# Create a DataFrame for predictions
predictions_df = pd.DataFrame({
    'Year': future_years.flatten(),
    'Predicted Financial Leverage': predicted_leverage
})

print(predictions_df)
# Simulate "actual" future financial leverage for evaluation
actual_future_leverage = np.random.rand(len(future_years)) * 0.1 + np.linspace(1.5, 2.5, len(future_years))

# Calculate and print RMSE (Root Mean Squared Error)
rmse = np.sqrt(mean_squared_error(actual_future_leverage, predicted_leverage))
print(f"RMSE: {rmse:.4f}")

# Calculate R² on historical data
r_squared = model.score(years, financial_leverage)
print(f"R² on historical data: {r_squared:.4f}")

# Simulate "actual" future financial leverage for evaluation (as before)
actual_future_leverage = np.random.rand(len(future_years)) * 0.1 + np.linspace(1.5, 2.5, len(future_years))

from sklearn.metrics import r2_score

# Calculate R² for future predictions
r_squared_future = r2_score(actual_future_leverage, predicted_leverage)
print(f"R² on future predictions: {r_squared_future:.4f}")

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression

# Simulated historical data
np.random.seed(0)  # For reproducible output
historical_years = np.arange(2010, 2023).reshape(-1, 1)
historical_leverage = np.random.rand(len(historical_years)) * 0.1 + np.linspace(0.5, 1.5, len(historical_years))

# Fit the linear regression model
model = LinearRegression()
model.fit(historical_years, historical_leverage)

# Future years to predict
future_years = np.arange(2023, 2036).reshape(-1, 1)
predicted_leverage = model.predict(future_years)

# Combine historical and future data
all_years = np.vstack((historical_years, future_years))
all_leverage = np.hstack((historical_leverage, predicted_leverage))

# Create DataFrame
df = pd.DataFrame({
    'Year': all_years.flatten(),
    'Financial Leverage': all_leverage
})

print(df)

def extract_leverage_data_by_bank(sheet_name, file_path):
    # Load the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Skip the 'En milliers de dinars' and 'Secteur' columns if they are not relevant
    relevant_columns = [col for col in df.columns if col not in ['En milliers de dinars', 'Secteur']]
    
    # Create a dictionary to hold your leverage data
    leverage_data = {}
    for bank_name in relevant_columns:
        # Assuming the leverage data you need is in the first row, adjust as necessary
        leverage_data[bank_name] = df[bank_name][0]
        
    return leverage_data

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt

# Assuming your existing function to extract leverage data remains the same.
# extract_leverage_data_by_bank(sheet_name, file_path)

# Load the Excel file
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)

# Create a dictionary to store data for each bank
data_by_bank = {}

# Loop through each sheet (year) and aggregate data by bank
for sheet_name in xl.sheet_names:
    if 'BILAN' in sheet_name:  # This checks if the sheet name contains 'BILAN'
        year = sheet_name.split()[-1]  # Extract the year from the sheet name
        leverage_data = extract_leverage_data_by_bank(sheet_name, file_path)
        for bank, leverage in leverage_data.items():
            if bank not in data_by_bank:
                data_by_bank[bank] = []
            data_by_bank[bank].append((year, leverage))

# Convert the aggregated data into a DataFrame for each bank and perform linear regression
for bank, data in data_by_bank.items():
    df = pd.DataFrame(data, columns=['Year', 'Financial Leverage']).dropna()
    df['Year'] = pd.to_numeric(df['Year'])
    
    if not df.empty:
        X = df[['Year']]
        y = df['Financial Leverage']
        model = LinearRegression().fit(X, y)
        
        # Predict future leverage
        future_years = np.arange(df['Year'].max() + 1, df['Year'].max() + 11).reshape(-1, 1)
        predicted_leverage = model.predict(future_years)
        
        # Plotting
        plt.figure(figsize=(10, 6))
        plt.plot(df['Year'], y, label='Historical Financial Leverage')
        plt.plot(future_years.flatten(), predicted_leverage, label='Predicted Financial Leverage', linestyle='--')
        plt.title(f'Financial Leverage Prediction for {bank}')
        plt.xlabel('Year')
        plt.ylabel('Financial Leverage')
        plt.legend()
        plt.show()

import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error

# Simulating historical financial leverage data for three banks
np.random.seed(0)  # For reproducible output
banks = ['ATB', 'BIAT', 'UBCI']
num_years = 23
data = []

# Générer des données de levier financier simulées pour chaque banque
for bank in banks:
    years = np.arange(2000, 2000 + num_years).reshape(-1, 1)
    financial_leverage = np.random.rand(num_years) * 0.1 + np.linspace(0.5, 1.5, num_years)
    for i in range(num_years):
        data.append([bank, years[i][0], financial_leverage[i]])

# Créer un DataFrame à partir des données
df = pd.DataFrame(data, columns=['Banque', 'Année', 'Levier Financier'])

# Initialiser les prédictions et les erreurs RMSE
predictions = []
rmses = []

# Entraîner un modèle de régression linéaire pour chaque banque et faire des prédictions
for bank in banks:
    bank_data = df[df['Banque'] == bank]
    model = LinearRegression()
    model.fit(bank_data[['Année']], bank_data['Levier Financier'])
    
    # Années à prédire
    future_years = np.arange(2023, 2023 + 10).reshape(-1, 1)
    predicted_leverage = model.predict(future_years)
    
    # Créer un DataFrame pour les prédictions pour chaque banque
    predictions_df = pd.DataFrame({
        'Année': future_years.flatten(),
        'Levier Financier Prévu': predicted_leverage
    })
    print(f"Prédictions pour {bank}:")
    print(predictions_df, "\n")
    
    # Ajouter les prédictions au tableau global des prédictions
    predictions.append((bank, predictions_df))
    
    # Simuler "réel" futur levier financier pour évaluation (pour exemple)
    actual_future_leverage = np.random.rand(len(future_years)) * 0.1 + np.linspace(1.5, 2.5, len(future_years))
    rmse = np.sqrt(mean_squared_error(actual_future_leverage, predicted_leverage))
    rmses.append((bank, rmse))

# Afficher l'erreur RMSE pour chaque banque
for bank, rmse in rmses:
    print(f"RMSE pour {bank}: {rmse:.4f}")


# Calculer R² pour chaque banque
for bank in banks:
    bank_data = df[df['Banque'] == bank]
    model = LinearRegression()
    X = bank_data[['Année']]
    y = bank_data['Levier Financier']
    model.fit(X, y)
    
    # Calculer R² sur les données historiques
    r_squared = model.score(X, y)
    print(f"R² pour {bank} sur les données historiques: {r_squared:.4f}")
# Pour chaque banque, ajuster un modèle, faire des prédictions et évaluer
for bank in banks:
    bank_data = df[df['Banque'] == bank]
    model = LinearRegression()
    X = bank_data[['Année']]
    y = bank_data['Levier Financier']
    model.fit(X, y)
    
    future_years = np.arange(2023, 2023 + 10).reshape(-1, 1)
    predicted_leverage = model.predict(future_years)
    
    # Simuler les valeurs réelles futures spécifiques à chaque banque pour l'évaluation
    np.random.seed(len(bank))  # Utiliser la longueur du nom de la banque comme graine pour la reproductibilité
    actual_future_leverage = np.random.rand(len(future_years)) * 0.1 + np.linspace(1.5, 2.5, len(future_years))
    
from sklearn.model_selection import cross_val_score

# Exemple d'utilisation de la validation croisée pour évaluer un modèle linéaire
model = LinearRegression()
scores = cross_val_score(model, X, y, cv=5, scoring='r2')

print(f"Scores de validation croisée (R²) : {scores}")
print(f"Moyenne des scores de validation croisée (R²) : {np.mean(scores):.4f}")

from sklearn.cluster import KMeans

# Convertir les données de levier financier en un tableau 2D pour K-Means
X = leverage_results_updated.values.reshape(-1, 1)

# Déterminer le nombre optimal de clusters en utilisant la méthode du coude
distortions = []
K = range(1, 10)
for k in K:
    kmeanModel = KMeans(n_clusters=k)
    kmeanModel.fit(X)
    distortions.append(kmeanModel.inertia_)

# Tracer le graphique du coude pour trouver le nombre optimal de clusters
plt.figure(figsize=(8, 6))
plt.plot(K, distortions, 'bx-')
plt.xlabel('Nombre de clusters')
plt.ylabel('Distortion')
plt.title('Méthode du coude pour le nombre optimal de clusters')
plt.show()

# Basé sur le graphique du coude, choisissons un nombre de clusters (par exemple, 3)
num_clusters = 3

# Appliquer K-Means avec le nombre de clusters choisi
kmeans = KMeans(n_clusters=num_clusters)
kmeans.fit(X)

# Ajouter les étiquettes de cluster aux données de levier financier
leverage_results_updated['Cluster'] = kmeans.labels_

# Afficher les résultats du clustering
print("Résultats du clustering basé sur l'effet de levier financier :")
print(leverage_results_updated)

# Plotting the clusters
plt.figure(figsize=(10, 6))
for cluster in range(num_clusters):
    cluster_data = leverage_results_updated[leverage_results_updated['Cluster'] == cluster]
    plt.scatter(cluster_data.index, cluster_data['Financial Leverage'], label=f'Cluster {cluster+1}')

plt.title('Clustering des années en fonction de l\'effet de levier financier')
plt.xlabel('Année')
plt.ylabel('Effet de levier financier')
plt.xticks(range(2010, 2023))
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score

# Actual evaluation metrics from the code snippets you provided
rmse_lr = np.sqrt(mean_squared_error(y_test, y_pred))
r2_lr = r2_score(y_test, y_pred)

rmse_rf = np.sqrt(mean_squared_error(y_test, y_pred_rf))
r2_rf = r2_score(y_test, y_pred_rf)

rmse_xgb = np.sqrt(mean_squared_error(y_test, y_pred_xgb))
r2_xgb = r2_score(y_test, y_pred_xgb)

# Assuming ARIMA RMSE and R² are calculated similarly
# Please replace `y_arima_pred` with your actual ARIMA predictions
rmse_arima = np.sqrt(mean_squared_error(leverage_results_updated['Financial Leverage'], in_sample_forecast))
r2_arima = r2_score(leverage_results_updated['Financial Leverage'], in_sample_forecast)

# Data preparation
data = {
    'Method': ['Linear Regression', 'Random Forest', 'XGBoost', 'ARIMA'],
    'RMSE': [rmse_lr, rmse_rf, rmse_xgb, rmse_arima],
    'R2': [r2_lr, r2_rf, r2_xgb, r2_arima]
}

df = pd.DataFrame(data)

# Plotting
fig, ax1 = plt.subplots(figsize=(10, 6))

# Bar plot for RMSE
ax1.set_xlabel('Method')
ax1.set_ylabel('RMSE', color='tab:red')
ax1.bar(df['Method'], df['RMSE'], color='tab:red', width=0.4, label='RMSE')
ax1.tick_params(axis='y', labelcolor='tab:red')

# Create a twin Axes sharing the xaxis for R²
ax2 = ax1.twinx()
ax2.set_ylabel('R²', color='tab:blue')
ax2.plot(df['Method'], df['R2'], color='tab:blue', marker='o', label='R²')
ax2.tick_params(axis='y', labelcolor='tab:blue')

# Title and legend
fig.tight_layout()
plt.title('Comparison of Evaluation Methods Based on Real Values')
fig.legend(loc="upper right", bbox_to_anchor=(1,1), bbox_transform=ax1.transAxes)

plt.show()

import pandas as pd

# Chemin d'accès au fichier Excel ajusté pour l'environnement actuel
file_path= 'EtudeSB - Copie.xlsx'
xl = pd.ExcelFile(file_path)

def calculate_ratios_for_year_adapted(bilan_df):
    try:
        # Sélectionner la dernière colonne pour les calculs, qui correspond aux totaux du secteur
        column_for_calculation = bilan_df.iloc[:, -1]
        
        # Correspondance des termes et somme des valeurs pour les postes spécifiques
        liquidites = bilan_df[bilan_df['En milliers de dinars'].str.contains('Caisse et avoirs auprès de la BCT CCP et TGT', na=False)]['Secteur'].sum()
        creances_sur_clientele = bilan_df[bilan_df['En milliers de dinars'].str.contains('Créances sur clientèle', na=False)]['Secteur'].sum()
        creances_financieres = bilan_df[bilan_df['En milliers de dinars'].str.contains('Créances sur établissement financier ou bancaires', na=False)]['Secteur'].sum()
        depots_et_avoirs_clients = bilan_df[bilan_df['En milliers de dinars'].str.contains('Dépôts et avoirs de la clientèle', na=False)]['Secteur'].sum()
        emprunts_et_ressources_speciales = bilan_df[bilan_df['En milliers de dinars'].str.contains('Emprunts et ressources spéciales', na=False)]['Secteur'].sum()
        autres_passifs = bilan_df[bilan_df['En milliers de dinars'].str.contains('Autres passifs', na=False)]['Secteur'].sum()

        # Calcul du passif à court terme
        passifs_court_terme = depots_et_avoirs_clients + emprunts_et_ressources_speciales + autres_passifs

        # Calcul du ratio de liquidité
        if passifs_court_terme == 0:
            ratio_liquidite = float('nan')  # Ou définir à 0 selon la préférence
        else:
            ratio_liquidite = ((liquidites + creances_sur_clientele + creances_financieres) / passifs_court_terme) * 100

    except Exception as e:
        print(f"Une erreur est survenue lors du traitement des données : {e}")
        return (0, 0, 0, 0, 0)  # Retourne des valeurs par défaut en cas d'erreur

    return (liquidites, creances_sur_clientele, creances_financieres, passifs_court_terme, ratio_liquidite)

# Initialiser un dictionnaire pour contenir les données collectées pour chaque année
yearly_data_adapted = {
    'Year': [],
    'Liquidites': [],
    'Creances_sur_Clientele': [],
    'Creances_Financieres': [],
    'Passifs_Court_Terme': [],
    'Ratio_Liquidite (%)': []
}

# Itérer sur chaque feuille de bilan de 2010 à 2022
for year in range(2010, 2023):
    sheet_name = f'BILAN {year}'
    try:
        bilan_df = xl.parse(sheet_name)
    except ValueError as ve:
        print(f"Erreur lors de la lecture de la feuille {sheet_name}: {ve}")
        continue

    liquidites, creances_sur_clientele, creances_financieres, passifs_ct, ratio_liquidite = calculate_ratios_for_year_adapted(bilan_df)
    
    yearly_data_adapted['Year'].append(year)
    yearly_data_adapted['Liquidites'].append(liquidites)
    yearly_data_adapted['Creances_sur_Clientele'].append(creances_sur_clientele)
    yearly_data_adapted['Creances_Financieres'].append(creances_financieres)
    yearly_data_adapted['Passifs_Court_Terme'].append(passifs_ct)
    yearly_data_adapted['Ratio_Liquidite (%)'].append(ratio_liquidite)
# Spécifier le chemin d'accès et le nom du fichier où vous souhaitez enregistrer le DataFrame
output_file_path = 'Yearly_Ratios_Summary.xlsx'

# Enregistrer le DataFrame sous forme de fichier Excel

# Convertir le dictionnaire adapté en DataFrame pour une visualisation et une analyse faciles
yearly_summary_df_adapted = pd.DataFrame(yearly_data_adapted)
yearly_summary_df_adapted.to_excel(output_file_path, sheet_name='Ratios Summary', index=False)

# Afficher le résumé des données adaptées pour vérifier les résultats, particulièrement pour 2021
yearly_summary_df_adapted

import plotly.express as px

# Créer un graphique interactif pour visualiser l'évolution du Ratio de Liquidité
fig = px.line(yearly_summary_df_adapted, x='Year', y='Ratio_Liquidite (%)', title='Évolution du Ratio de Liquidité de 2010 à 2022', markers=True)

# Ajouter des détails au graphique
fig.update_layout(xaxis_title='Année', yaxis_title='Ratio de Liquidité (%)', legend_title='Légende')

# Afficher le graphique
fig.show()

import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.metrics import mean_squared_error, r2_score, silhouette_score
import numpy as np
import matplotlib.pyplot as plt
from scipy.spatial.distance import cdist
# Supposons que 'Year' a été remplacé par 'Bank' juste pour l'exemple
# Assurez-vous que 'Bank' est la colonne correcte à utiliser
# Sélection de la colonne numérique comme variable explicative
X = yearly_summary_df_adapted[['Liquidites', 'Creances_Financieres']].values
y = yearly_summary_df_adapted['Ratio_Liquidite (%)'].values  
# Assurez-vous d'abord que 'bank_summary_df' est correctement défini et contient les données attendues
# Si nécessaire, remplacez 'bank_summary_df' par le nom correct de votre DataFrame contenant les données

# Variable cible

# Création et entraînement du modèle de régression linéaire
model = LinearRegression()
model.fit(X, y)

# Prédictions
predictions = model.predict(X)

# Évaluation du modèle
print("Mean squared error: %.2f" % mean_squared_error(y, predictions))
print('R²: %.2f' % r2_score(y, predictions))
# Assuming X is 2D, we plot each feature against y in a loop for simplicity. Adjust as necessary.

# Plotting the real values against the predictions
plt.figure(figsize=(10, 6))

# Plotting the real values
plt.scatter(X[:, 0], y, color='blue', label='Real Values: Liquidités')
plt.scatter(X[:, 1], y, color='green', label='Real Values: Créances Financières')

# Plotting the predictions
plt.plot(X[:, 0], predictions, color='red', linewidth=2, label='Predictions: Liquidités')
plt.plot(X[:, 1], predictions, color='orange', linewidth=2, label='Predictions: Créances Financières')

# Adding labels and title
plt.xlabel('Values')
plt.ylabel('Ratio de Liquidité (%)')
plt.title('Régression Linéaire: Réalité vs. Prédictions')
plt.legend()
plt.grid(True)

# Showing the plot
plt.show()

# Détermination du nombre optimal de clusters via la méthode du coude
distortions = []
K = range(1, 10)
for k in K:
    kmeanModel = KMeans(n_clusters=k)
    kmeanModel.fit(yearly_summary_df_adapted[['Ratio_Liquidite (%)']])
    distortions.append(sum(np.min(cdist(yearly_summary_df_adapted[['Ratio_Liquidite (%)']], kmeanModel.cluster_centers_, 'euclidean'), axis=1)) / yearly_summary_df_adapted.shape[0])

# Affichage de la méthode du coude
plt.figure(figsize=(16,8))
plt.plot(K, distortions, 'bx-')
plt.xlabel('k')
plt.ylabel('Distortion')
plt.title('La méthode du coude montrant le nombre optimal de clusters')
plt.show()

# Supposons que le nombre optimal de clusters déterminé soit 3
kmeans = KMeans(n_clusters=3)
kmeans.fit(yearly_summary_df_adapted[['Ratio_Liquidite (%)']])
labels = kmeans.labels_

# Évaluation avec le score de silhouette
silhouette_avg = silhouette_score(yearly_summary_df_adapted[['Ratio_Liquidite (%)']], labels)
print(f'Score de silhouette pour 3 clusters : {silhouette_avg}')

# Adapter le script pour calculer les ratios par banque à partir de la feuille "BILAN 2022"
# Lire les données de la feuille "BILAN 2022" pour examiner la structure
bilan_2022_df = xl.parse('BILAN 2022')

# Afficher les premières lignes pour comprendre la structure des données
bilan_2022_df.head()

# Initialiser un dictionnaire pour contenir les données collectées pour chaque banque
bank_data = {
    'Bank': [],
    'Liquidites': [],
    'Creances_sur_Clientele': [],
    'Creances_Financieres': [],
    'Passifs_Court_Terme': [],
    'Ratio_Liquidite (%)': []
}

# Itérer sur chaque colonne (banque) de la feuille de bilan, en excluant la première (descriptions) et la dernière (secteur total)
for bank in bilan_2022_df.columns[1:-1]:  # Exclure la première et la dernière colonne
    bilan_df = bilan_2022_df[['En milliers de dinars', bank]]
    
    # Renommer les colonnes pour simplifier l'accès
    bilan_df.columns = ['Postes', 'Valeurs']
    
    # Effectuer les calculs comme précédemment, mais pour chaque banque
    liquidites = bilan_df[bilan_df['Postes'].str.contains('Caisse et avoirs auprès de la BCT CCP et TGT', na=False)]['Valeurs'].sum()
    creances_sur_clientele = bilan_df[bilan_df['Postes'].str.contains('Créances sur clientèle', na=False)]['Valeurs'].sum()
    creances_financieres = bilan_df[bilan_df['Postes'].str.contains('Créances sur établissement financier ou bancaires', na=False)]['Valeurs'].sum()
    # Hypothèse: Utiliser les mêmes postes pour calculer les passifs à court terme
    depots_et_avoirs_clients = bilan_df[bilan_df['Postes'].str.contains('Dépôts et avoirs de la clientèle', na=False)]['Valeurs'].sum()
    emprunts_et_ressources_speciales = bilan_df[bilan_df['Postes'].str.contains('Emprunts et ressources spéciales', na=False)]['Valeurs'].sum()
    autres_passifs = bilan_df[bilan_df['Postes'].str.contains('Autres passifs', na=False)]['Valeurs'].sum()
    
    # Calcul du passif à court terme (sans indication claire, on utilise les mêmes postes hypothétiquement)
    passifs_court_terme = depots_et_avoirs_clients + emprunts_et_ressources_speciales + autres_passifs
    
    # Calcul du ratio de liquidité
    if passifs_court_terme == 0:
        ratio_liquidite = float('nan')  # Ou définir à 0 selon la préférence
    else:
        ratio_liquidite = ((liquidites + creances_sur_clientele + creances_financieres) / passifs_court_terme) * 100

    # Ajouter les résultats au dictionnaire
    bank_data['Bank'].append(bank)
    bank_data['Liquidites'].append(liquidites)
    bank_data['Creances_sur_Clientele'].append(creances_sur_clientele)
    bank_data['Creances_Financieres'].append(creances_financieres)
    bank_data['Passifs_Court_Terme'].append(passifs_court_terme)
    bank_data['Ratio_Liquidite (%)'].append(ratio_liquidite)
 #Spécifier le chemin d'accès et le nom du fichier où vous souhaitez enregistrer le DataFrame

# Enregistrer le DataFrame sous forme de fichier Excel
# Convertir le dictionnaire en DataFrame pour une visualisation et une analyse faciles
# Spécifier le chemin d'accès et le nom du fichier où vous souhaitez enregistrer le DataFrame
output_file_path = 'bANK_Ratios_Summary.xlsx'

# Convertir le dictionnaire en DataFrame pour une visualisation et une analyse faciles
bank_summary_df_adapted = pd.DataFrame(bank_data)

# Enregistrer le DataFrame sous forme de fichier Excel
bank_summary_df_adapted.to_excel(output_file_path, sheet_name='Ratios Summary', index=False)

# Convertir le dictionnaire adapté en DataFrame pour une visualisation et une analyse faciles

# Afficher le résumé des données pour vérifier les résultats
bank_summary_df_adapted.head()  # Afficher les premières lignes pour vérification

import plotly.express as px

# Créer un graphique à barres interactif pour visualiser le Ratio de Liquidité par banque
fig = px.bar(bank_summary_df_adapted, x='Bank', y='Ratio_Liquidite (%)', title='Ratio de Liquidité par Banque pour l\'année 2022', text='Ratio_Liquidite (%)')

# Ajouter des détails au graphique
fig.update_layout(xaxis_title='Banque', yaxis_title='Ratio de Liquidité (%)', legend_title='Légende')
fig.update_traces(texttemplate='%{text:.2s}', textposition='outside')

# Afficher le graphique
fig.show()

import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.metrics import mean_squared_error, r2_score, silhouette_score
import numpy as np
import matplotlib.pyplot as plt
from scipy.spatial.distance import cdist
# Supposons que 'Year' a été remplacé par 'Bank' juste pour l'exemple
# Assurez-vous que 'Bank' est la colonne correcte à utiliser
# Sélection de la colonne numérique comme variable explicative
X = bank_summary_df_adapted[['Liquidites', 'Creances_Financieres']].values
y = bank_summary_df_adapted['Ratio_Liquidite (%)'].values  

# Assurez-vous d'abord que 'bank_summary_df' est correctement défini et contient les données attendues
# Si nécessaire, remplacez 'bank_summary_df' par le nom correct de votre DataFrame contenant les données

# Variable cible

# Création et entraînement du modèle de régression linéaire
model = LinearRegression()
model.fit(X, y)

# Prédictions
predictions = model.predict(X)

# Évaluation du modèle
print("Mean squared error: %.2f" % mean_squared_error(y, predictions))
print('R²: %.2f' % r2_score(y, predictions))


# Plotting the real values against the predictions
plt.figure(figsize=(10, 6))

# Plotting the real values
plt.scatter(X[:, 0], y, color='blue', label='Real Values: Liquidités')
plt.scatter(X[:, 1], y, color='green', label='Real Values: Créances Financières')

# Plotting the predictions
plt.plot(X[:, 0], predictions, color='red', linewidth=2, label='Predictions: Liquidités')
plt.plot(X[:, 1], predictions, color='orange', linewidth=2, label='Predictions: Créances Financières')

# Adding labels and title
plt.xlabel('Values')
plt.ylabel('Ratio de Liquidité (%)')
plt.title('Régression Linéaire: Réalité vs. Prédictions')
plt.legend()
plt.grid(True)

# Showing the plot
plt.show()

# Détermination du nombre optimal de clusters via la méthode du coude
distortions = []
K = range(1, 10)
for k in K:
    kmeanModel = KMeans(n_clusters=k)
    kmeanModel.fit(bank_summary_df_adapted[['Ratio_Liquidite (%)']])
    distortions.append(sum(np.min(cdist(yearly_summary_df_adapted[['Ratio_Liquidite (%)']], kmeanModel.cluster_centers_, 'euclidean'), axis=1)) / yearly_summary_df_adapted.shape[0])

# Affichage de la méthode du coude
plt.figure(figsize=(16,8))
plt.plot(K, distortions, 'bx-')
plt.xlabel('k')
plt.ylabel('Distortion')
plt.title('La méthode du coude montrant le nombre optimal de clusters')
plt.show()

# Supposons que le nombre optimal de clusters déterminé soit 3
kmeans = KMeans(n_clusters=3)
kmeans.fit(bank_summary_df_adapted[['Ratio_Liquidite (%)']])
labels = kmeans.labels_

# Évaluation avec le score de silhouette
silhouette_avg = silhouette_score(bank_summary_df_adapted[['Ratio_Liquidite (%)']], labels)
print(f'Score de silhouette pour 3 clusters : {silhouette_avg}')

import pandas as pd

# Assurez-vous que xl est votre objet ExcelFile chargé avec le fichier Excel

# Initialiser un dictionnaire pour contenir les données collectées pour chaque année et chaque banque
all_bank_data = {
    'Year': [],
    'Bank': [],
    'Liquidites': [],
    'Creances_sur_Clientele': [],
    'Creances_Financieres': [],
    'Passifs_Court_Terme': [],
    'Ratio_Liquidite (%)': []
}

# Boucler sur chaque année de 2010 à 2023
for year in range(2010, 2023):
    sheet_name = f'BILAN {year}'
    try:
        bilan_df = xl.parse(sheet_name)
    except ValueError as ve:
        print(f"Erreur lors de la lecture de la feuille {sheet_name}: {ve}")
        continue

    # Itérer sur chaque colonne (banque)
    for bank in bilan_df.columns[1:-1]:  # Exclure la première et la dernière colonne
        bilan_bank_df = bilan_df[['En milliers de dinars', bank]]
        bilan_bank_df.columns = ['Postes', 'Valeurs']
        
        # Convertir les valeurs en nombres flottants
        bilan_bank_df['Valeurs'] = pd.to_numeric(bilan_bank_df['Valeurs'], errors='coerce').fillna(0)
        
        liquidites = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Caisse et avoirs auprès de la BCT CCP et TGT', na=False)]['Valeurs'].sum()
        creances_sur_clientele = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Créances sur clientèle', na=False)]['Valeurs'].sum()
        creances_financieres = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Créances sur établissement financier ou bancaires', na=False)]['Valeurs'].sum()
        depots_et_avoirs_clients = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Dépôts et avoirs de la clientèle', na=False)]['Valeurs'].sum()
        emprunts_et_ressources_speciales = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Emprunts et ressources spéciales', na=False)]['Valeurs'].sum()
        autres_passifs = bilan_bank_df[bilan_bank_df['Postes'].str.contains('Autres passifs', na=False)]['Valeurs'].sum()

        # Calcul du passif à court terme
        passifs_court_terme = depots_et_avoirs_clients + emprunts_et_ressources_speciales + autres_passifs

        # Calcul du ratio de liquidité
        if passifs_court_terme == 0:
            ratio_liquidite = float('nan')  # Ou définir à 0 selon la préférence
        else:
            ratio_liquidite = ((liquidites + creances_sur_clientele + creances_financieres) / passifs_court_terme) * 100

        # Ajouter les résultats au dictionnaire
        all_bank_data['Year'].append(year)
        all_bank_data['Bank'].append(bank)
        all_bank_data['Liquidites'].append(liquidites)
        all_bank_data['Creances_sur_Clientele'].append(creances_sur_clientele)
        all_bank_data['Creances_Financieres'].append(creances_financieres)
        all_bank_data['Passifs_Court_Terme'].append(passifs_court_terme)
        all_bank_data['Ratio_Liquidite (%)'].append(ratio_liquidite)
 #Spécifier le chemin d'accès et le nom du fichier où vous souhaitez enregistrer le DataFrame
output_file_path = 'YearlyBank_Ratios_Summary.xlsx'

# Enregistrer le DataFrame sous forme de fichier Excel
yearly_summary_df_adapted.to_excel(output_file_path, sheet_name='Ratios Summary', index=False)
# Convertir le dictionnaire en DataFrame pour une visualisation et une analyse faciles
banka_summary_df = pd.DataFrame(all_bank_data)

# Afficher le résumé des données pour vérifier les résultats
print(banka_summary_df.head())  # Afficher les premières lignes pour vérification

import plotly.express as px

# Création du graphique
fig = px.line(banka_summary_df,
              x='Year', y='Ratio_Liquidite (%)',
              color='Bank',
              title='Évolution du Ratio de Liquidité par Banque de 2010 à 2022',
              markers=True)

# Ajout de détails au graphique
fig.update_layout(xaxis_title='Année',
                  yaxis_title='Ratio de Liquidité (%)',
                  legend_title='Banque')

# Affichage du graphique
fig.show()

from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
import pandas as pd
from sklearn.impute import SimpleImputer

# Préparation des données financières avec imputation des valeurs NaN
imputer = SimpleImputer(strategy='mean')  # Création d'un imputeur qui remplace les valeurs NaN par la moyenne de chaque colonne
imputed_features = imputer.fit_transform(banka_summary_df[['Liquidites', 'Creances_sur_Clientele', 'Creances_Financieres', 'Passifs_Court_Terme', 'Ratio_Liquidite (%)']])

# Standardisation des données financières après imputation
scaler = StandardScaler()
scaled_features = scaler.fit_transform(imputed_features)

# Application de l'analyse de cluster K-Means
kmeans = KMeans(n_clusters=3, random_state=42)  # Ajustez le nombre de clusters si nécessaire
banka_summary_df['Cluster'] = kmeans.fit_predict(scaled_features)
#Spécifier le chemin d'accès et le nom du fichier où vous souhaitez enregistrer le DataFrame
output_file_path = 'cluster.xlsx'

# Enregistrer le DataFrame sous forme de fichier Excel
yearly_summary_df_adapted.to_excel(output_file_path, sheet_name='Ratios Summary', index=False)
# Affichage des premières lignes pour vérifier les clusters
print(banka_summary_df.head())

from sklearn.metrics import silhouette_score

# Assurez-vous que 'clusters' contient les étiquettes de cluster pour chaque banque
clusters = banka_summary_df['Cluster']

# Calcul du score de silhouette pour évaluer la qualité des clusters
silhouette_avg = silhouette_score(scaled_features, clusters)
print(f"Score de silhouette moyen : {silhouette_avg:.2f}")

import plotly.express as px

# Création d'un graphique à barres pour visualiser le Ratio de Liquidité par Banque
fig = px.bar(banka_summary_df, x='Bank', y='Ratio_Liquidite (%)', color='Cluster', text='Ratio_Liquidite (%)', title="Ratio de Liquidité par Banque")

# Ajout des détails au graphique
fig.update_layout(xaxis_title='Banque', yaxis_title='Ratio de Liquidité (%)', legend_title='Cluster')
fig.update_traces(texttemplate='%{text:.2s}', textposition='outside')

# Pour afficher le nom des banques sur le graphique, ajustez la disposition des axes et des étiquettes
fig.update_xaxes(tickangle=-45)

# Affichage du graphique
fig.show()

from sklearn.decomposition import PCA

# Réduction de la dimensionnalité
pca = PCA(n_components=2)  # Réduire à 2 dimensions pour la visualisation
pca_features = pca.fit_transform(scaled_features)

# Ajouter les composantes principales comme nouvelles colonnes pour la visualisation
banka_summary_df['PCA1'] = pca_features[:, 0]
banka_summary_df['PCA2'] = pca_features[:, 1]

import plotly.express as px

# Création du graphique des clusters
fig = px.scatter(banka_summary_df, x='PCA1', y='PCA2',
                 color='Cluster',
                 title="Visualisation des Clusters des Banques",
                 labels={'PCA1': 'Composante Principale 1', 'PCA2': 'Composante Principale 2'},
                 hover_data=['Bank'])  # Afficher le nom de la banque au survol

# Ajouter des détails au graphique
fig.update_layout(legend_title_text='Cluster')
fig.show()


import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.metrics import mean_squared_error, r2_score, silhouette_score
import numpy as np
import matplotlib.pyplot as plt
from scipy.spatial.distance import cdist

# Préparation des données pour la régression linéaire
X = yearly_summary_df_adapted[['Year']].values.reshape(-1, 1)  # Feature matrix
y = yearly_summary_df_adapted['Ratio_Liquidite (%)'].values  # Target variable

# Entraînement du modèle de régression linéaire
model = LinearRegression()
model.fit(X, y)

# Prédictions
predictions = model.predict(X)

# Évaluation du modèle de régression linéaire
r2 = r2_score(y, predictions)
rmse = np.sqrt(mean_squared_error(y, predictions))

print(f"R² score: {r2}")
print(f"RMSE: {rmse}")

import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score
import numpy as np

# Supposons que 'Year' a été remplacé par 'Bank' juste pour l'exemple
# Assurez-vous que 'Bank' est la colonne correcte à utiliser

# Sélection de la colonne numérique comme variable explicative
X = yearly_summary_df_adapted[['Liquidites', 'Creances_Financieres']].values
y = yearly_summary_df_adapted['Ratio_Liquidite (%)'].values  

# Création et entraînement du modèle de régression linéaire
linear_model = LinearRegression()
linear_model.fit(X, y)

# Prédictions avec le modèle de régression linéaire
linear_predictions = linear_model.predict(X)

# Calcul des métriques d'évaluation pour le modèle de régression linéaire
linear_mse = mean_squared_error(y, linear_predictions)
linear_r2 = r2_score(y, linear_predictions)

# Création et entraînement du modèle de forêt aléatoire
rf_model = RandomForestRegressor()
rf_model.fit(X, y)

# Prédictions avec le modèle de forêt aléatoire
rf_predictions = rf_model.predict(X)

# Calcul des métriques d'évaluation pour le modèle de forêt aléatoire
rf_mse = mean_squared_error(y, rf_predictions)
rf_r2 = r2_score(y, rf_predictions)

# Affichage des résultats
print("Métriques d'évaluation pour la régression linéaire :")
print("MSE :", linear_mse)
print("R² :", linear_r2)
print("\nMétriques d'évaluation pour le modèle de forêt aléatoire :")
print("MSE :", rf_mse)
print("R² :", rf_r2)

import matplotlib.pyplot as plt

# Plotting the real values against the predictions for both Linear Regression and Random Forest
plt.figure(figsize=(10, 6))

# Plotting Linear Regression
plt.scatter(y, linear_predictions, color='blue', label='Linear Regression')

# Plotting Random Forest
plt.scatter(y, rf_predictions, color='green', label='Random Forest')

# Adding labels and title
plt.xlabel('Valeurs Réelles')
plt.ylabel('Prédictions')
plt.title('Comparaison des Prédictions - Régression Linéaire vs. Forêt Aléatoire')
plt.legend()
plt.grid(True)

# Showing the plot
plt.show()

import matplotlib.pyplot as plt

# Définir la largeur des barres d'histogramme
bar_width = 0.35

# Positions des barres pour les deux ensembles de données
index = np.arange(len(y))

# Créer une figure et des axes
plt.figure(figsize=(10, 6))

# Tracer l'histogramme des valeurs réelles
plt.bar(index, y, bar_width, color='blue', label='Valeurs Réelles')

# Tracer l'histogramme des prédictions de la régression linéaire
plt.bar(index + bar_width, linear_predictions, bar_width, color='green', label='Prédictions Régression Linéaire')

# Tracer l'histogramme des prédictions de la forêt aléatoire
plt.bar(index + 2*bar_width, rf_predictions, bar_width, color='red', label='Prédictions Forêt Aléatoire')

# Ajouter des étiquettes sur l'axe x
plt.xlabel('Échantillons')
plt.ylabel('Valeurs')
plt.title('Comparaison des Prédictions - Régression Linéaire vs. Forêt Aléatoire')
plt.xticks(index + bar_width, index)
plt.legend()

# Afficher le graphique
plt.tight_layout()
plt.show()
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import plotly.express as px
import plotly.graph_objects as go
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline
from statsmodels.tsa.arima.model import ARIMA
import statsmodels.api as sm
import pmdarima as pm
import os
from datetime import datetime

print("Bibliothèques importées avec succès.")

# Chemin d'accès au fichier Excel
excel_path = 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(excel_path)

# Affichage des noms des feuilles pour comprendre la structure du classeur
sheet_names = xls.sheet_names
print(f"Feuilles disponibles dans le fichier: {sheet_names}")

# Chargement d'une feuille spécifique pour examiner sa structure
balance_sheet_2022 = pd.read_excel(excel_path, sheet_name='BILAN 2022')
balance_sheet_2022.head()

# Nettoyage des données en supprimant les valeurs négatives et NaN pour la colonne 'Secteur'
balance_sheet_2022['Secteur'] = pd.to_numeric(balance_sheet_2022['Secteur'], errors='coerce').fillna(0)
cleaned_data = balance_sheet_2022[balance_sheet_2022['Secteur'] >= 0]
cleaned_data.head()

# Exemple: Affichage de statistiques descriptives de la colonne 'Secteur'
cleaned_data['Secteur'].describe()

# Exemple de visualisation: Distribution des actifs bancaires
plt.figure(figsize=(10, 6))
plt.hist(cleaned_data['Secteur'], bins=20, color='skyblue')
plt.title('Distribution des Actifs Bancaires en 2022')
plt.xlabel('Actifs')
plt.ylabel('Fréquence')
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px

# Initialize a list to hold aggregated data
aggregated_data_list = []

# Load and aggregate data from each year
for year in range(2010, 2023):  # From 2010 to 2022
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name).dropna(subset=['Secteur'])
    # Summarize total assets for the sector
    total_assets = df['Secteur'].sum()
    # Append the total to the aggregated data list
    aggregated_data_list.append({'Year': year, 'Total Assets': total_assets})

# Create DataFrame from the list of dictionaries
# Create DataFrame from the list of dictionaries
aggregated_data = pd.DataFrame(aggregated_data_list)

# Set 'Year' as the index
aggregated_data.set_index('Year', inplace=True)

# Utilisation de Plotly pour créer un graphique filtrable dynamique
fig = px.line(aggregated_data, x=aggregated_data.index, y='Total Assets',
              title='Total Banking Sector Assets in Tunisia (2010-2022)',
              labels={'x': 'Year', 'Total Assets': 'Total Assets (in thousands of dinars)'})

# Ajouter un slider et des boutons pour filtrer dynamiquement l'année
fig.update_layout(xaxis=dict(rangeselector=dict(buttons=list([
    dict(count=1, label='1y', step='year', stepmode='backward'),
    dict(count=2, label='2y', step='year', stepmode='backward'),
    dict(count=5, label='5y', step='year', stepmode='backward'),
    dict(step='all')
])),
    rangeslider=dict(visible=True),
    type='date'
))
import pandas as pd
import numpy as np
import plotly.graph_objects as go

# Chemin d'accès à votre fichier Excel
excel_path = 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(excel_path)

# Liste des banques sans doublons
banks = ['ATB', 'BNA', 'ATTIBK', 'BT', 'BTL', 'BTS', 'ABC', 'WIB', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB', 'BARAKA', 'BH', 'BTK', 'TSB', 'QNB', 'BTE']

# Initialiser un dictionnaire pour stocker les données pour chaque banque
bank_data = {bank: {'Years': [], 'Total Assets': []} for bank in banks}

# Années à itérer
years = range(2010, 2023)

# Itérer sur chaque année et extraire les données pour chaque banque
for year in years:
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    for bank in banks:
        if bank in df.columns:
            total_assets = pd.to_numeric(df[bank], errors='coerce').fillna(0).sum()
        else:
            total_assets = np.nan  # Utiliser np.nan pour les années où les données de la banque sont absentes

        # Stocker les données dans le dictionnaire
        bank_data[bank]['Years'].append(year)
        bank_data[bank]['Total Assets'].append(total_assets)

# Créer un graphique pour chaque banque
for bank in banks:
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=bank_data[bank]['Years'], y=bank_data[bank]['Total Assets'], mode='lines+markers', name=bank))
    
    fig.update_layout(title=f'Total Assets Trend for {bank} (2010-2022)',
                      xaxis_title='Year',
                      yaxis_title='Total Assets',
                      legend_title='Bank')
    
    fig.show()

import pandas as pd
import numpy as np
import plotly.graph_objects as go

# Chemin d'accès à votre fichier Excel - à ajuster selon votre environnement
excel_path = 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(excel_path)

# Liste des banques sans doublons
banks = ['ATB', 'BNA', 'ATTIBK', 'BT', 'BTS', 'ABC', 'WIB', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB', 'BARAKA', 'BH', 'BTK', 'QNB', 'BTE']

# Définir les années pour itérer
years = range(2010, 2023)

# Initialiser un DataFrame pour stocker les données de chaque banque
bank_data_df = pd.DataFrame(index=years)

# Itérer sur chaque année et extraire les données pour chaque banque
for year in years:
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    for bank in banks:
        if bank in df.columns:
            total_assets = pd.to_numeric(df[bank], errors='coerce').fillna(0).sum()
            bank_data_df.loc[year, bank] = total_assets
        else:
            bank_data_df.loc[year, bank] = np.nan

# Création du graphique interactif avec Plotly
fig = go.Figure()

for bank in banks:
    # Ici, bank_data_df.index est utilisé pour l'axe des x
    fig.add_trace(go.Scatter(x=bank_data_df.index, y=bank_data_df[bank], mode='lines+markers', name=bank))

fig.update_layout(title='Total Assets Trend for Banks (2010-2022)',
                  xaxis_title='Year',
                  yaxis_title='Total Assets',
                  legend_title='Bank')

import pandas as pd
import matplotlib.pyplot as plt

# Chemin d'accès au fichier Excel original
excel_path = 'EtudeSB - Copie.xlsx'
# Chemin d'accès pour sauvegarder le nouveau fichier Excel
output_file_path = 'dat.xlsx'

# Initialiser une liste pour stocker les données agrégées
aggregated_data_list = []

# Charger et agréger les données de chaque année
for year in range(2010, 2023):  # De 2010 à 2022
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name).dropna(subset=['Secteur'])
    # Résumer les actifs totaux pour le secteur
    total_assets = df['Secteur'].sum()
    # Ajouter le total à la liste des données agrégées
    aggregated_data_list.append({'Year': year, 'Total Assets': total_assets})

# Créer un DataFrame à partir de la liste de dictionnaires
aggregated_data = pd.DataFrame(aggregated_data_list)

# Définir 'Year' comme l'index
aggregated_data.set_index('Year', inplace=True)

# Affichage sous forme de tableau (dans la console ou le notebook Jupyter)
print(aggregated_data)

# Sauvegarder le DataFrame dans un fichier Excel
aggregated_data.to_excel(output_file_path)

# Graphique
plt.figure(figsize=(12, 6))
plt.plot(aggregated_data.index, aggregated_data['Total Assets'], marker='o', linestyle='-', color='b')
plt.title('Total des actifs du secteur bancaire en Tunisie (2010-2022)')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux (en milliers de dinars)')
plt.grid(True)
plt.xticks(aggregated_data.index, rotation=45)  # Rotation des étiquettes d'année pour une meilleure lisibilité
plt.tight_layout()  # Ajustement de la mise en page

# Afficher le graph

import pandas as pd
import matplotlib.pyplot as plt

# Chemin d'accès au fichier Excel
excel_path = 'EtudeSB - Copie.xlsx'

# Initialiser un dictionnaire pour stocker les données agrégées par banque
aggregated_data_by_bank = {}

# Charger et agréger les données de chaque année
for year in range(2010, 2023):  # De 2010 à 2022
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    # Itérer sur chaque colonne (banque) dans le DataFrame
    for bank in df.columns[1:]:  # Supposons que la première colonne est 'Secteur' et les autres sont les banques
        # Convertir les valeurs de la colonne banque en numérique, en ignorant les erreurs et en remplaçant les non numériques par NaN
        assets_numeric = pd.to_numeric(df[bank], errors='coerce')
        # Calculer la somme des actifs pour l'année en cours, en excluant les NaN
        total_assets_year = assets_numeric.sum()
        # Vérifier si la colonne (banque) existe déjà dans le dictionnaire
        if bank not in aggregated_data_by_bank:
            aggregated_data_by_bank[bank] = 0
        # Ajouter les actifs de l'année courante à la banque correspondante
        aggregated_data_by_bank[bank] += total_assets_year

# Conversion du dictionnaire en DataFrame pour une manipulation plus aisée
aggregated_data_df = pd.DataFrame(list(aggregated_data_by_bank.items()), columns=['Bank', 'Total Assets'])

# Tri des données par 'Total Assets' pour une meilleure visualisation
aggregated_data_df = aggregated_data_df.sort_values(by='Total Assets', ascending=False)
aggregated_data.to_excel(output_file_path)

# Graphique
plt.figure(figsize=(12, 8))
plt.bar(aggregated_data_df['Bank'], aggregated_data_df['Total Assets'], color='skyblue')
plt.title('Distribution des Actifs par Banque dans le Secteur Bancaire Tunisien (2010-2022)')
plt.xlabel('Banque')
plt.ylabel('Actifs Totaux')  # Modification ici pour enlever la référence spécifique à "en milliers de dinars"
plt.xticks(rotation=45, ha='right')  # Rotation des noms des banques pour une meilleure lisibilité
plt.tight_layout()

# Affichage du graphique
plt.show()

# Affichage du DataFrame
print(aggregated_data_df)
output_file_path = 'datBanques.xlsx'
aggregated_data_df.to_excel(output_file_path, index=False)


import matplotlib.pyplot as plt
import pandas as pd

# Chemin d'accès au fichier Excel - ajustez selon votre environnement
excel_path = 'EtudeSB - Copie.xlsx'

# Charger et agréger les données de chaque année
aggregated_data_list = []
for year in range(2010, 2023):  # De 2010 à 2022
    sheet_name = f'BILAN {year}'
    df = pd.read_excel(excel_path, sheet_name=sheet_name).dropna(subset=['Secteur'])
    total_assets = df['Secteur'].sum()
    aggregated_data_list.append({'Year': year, 'Total Assets': total_assets})

aggregated_data = pd.DataFrame(aggregated_data_list)
aggregated_data.set_index('Year', inplace=True)

# Visualisation de la tendance des actifs totaux
plt.figure(figsize=(10, 6))
plt.plot(aggregated_data.index, aggregated_data['Total Assets'], marker='o')
plt.title('Evolution des Actifs Totaux du Secteur Bancaire Tunisien (2010-2022)')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux')
plt.grid(True)
plt.show()

from sklearn.linear_model import LinearRegression
import numpy as np
import plotly.express as px

# Prepare the data for modeling
X = aggregated_data.index.values.reshape(-1, 1)  # Reshape years for sklearn
y = aggregated_data['Total Assets'].values
output_file_path = 'PredictionBA.xlsx'

# Create and fit the model
model = LinearRegression()
model.fit(X, y)

# Predict for future years, e.g., 2023 to 2025
future_years = np.array([[2023], [2024], [2025],[2026], [2027], [2028],[2029], [2030], [2031],[2032],[2033],[2034],[2035]])
predictions = model.predict(future_years)
aggregated_data.to_excel(output_file_path)

# Add predictions to the plot
for i, year in enumerate(future_years.flatten()):
    aggregated_data.loc[year] = predictions[i]

# Update Plotly figure to include predictions
fig = px.line(aggregated_data, x=aggregated_data.index, y='Total Assets',
              title='Total Banking Sector Assets in Tunisia (2010-2025 Predicted)',
              labels={'x': 'Year', 'Total Assets': 'Total Assets (in thousands of dinars)'})
# (Reconfigure the slider and buttons as needed to include the new years)

fig.show()
# Création d'un DataFrame pour les prédictions
predictionsDA_df = pd.DataFrame({
    'Year': future_years.flatten(),
    'Predicted Total Assets': predictions
})

# Affichage du DataFrame
print(predictionsDA_df)
output_file_path = 'PredictionAnnée.xlsx'
predictionsDA_df.to_excel(output_file_path, index=False)
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline

# Chemin d'accès au fichier Excel
excel_path = 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(excel_path)

# Charger les données agrégées
aggregated_data_path = 'dat.xlsx'
aggregated_data = pd.read_excel(aggregated_data_path, index_col='Year')

# Visualisation des actifs totaux sur la période étudiée
plt.figure(figsize=(12, 6))
plt.plot(aggregated_data.index, aggregated_data['Total Assets'], marker='o', linestyle='-', color='b')
plt.title('Total des actifs du secteur bancaire en Tunisie (2010-2022)')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux (en milliers de dinars)')
plt.grid(True)
plt.xticks(aggregated_data.index, rotation=45)
plt.tight_layout()
plt.show()

# Préparation des données pour le modèle
X = aggregated_data.index.values.reshape(-1, 1)  # Remodeler les années pour sklearn
y = aggregated_data['Total Assets'].values

# Créer et ajuster le modèle de régression polynomiale de degré 2
model = make_pipeline(PolynomialFeatures(degree=2), LinearRegression())
model.fit(X, y)

# Prédire pour les années futures
future_years = np.array([[2023], [2024], [2025], [2026], [2027], [2028], [2029], [2030], [2031], [2032], [2033], [2034], [2035]])
predictions = model.predict(future_years)

# Ajouter les prédictions au DataFrame
aggregated_data_future = pd.DataFrame(index=future_years.flatten(), columns=['Predicted Total Assets'])
aggregated_data_future['Predicted Total Assets'] = predictions

# Concaténer les données prédites avec les données existantes
aggregated_data_combined = pd.concat([aggregated_data, aggregated_data_future])

# Afficher les prédictions
print("Prédictions pour les années 2023-2032:")
print(aggregated_data_future)

# Mettre à jour le graphique
plt.figure(figsize=(12, 6))
plt.plot(aggregated_data_combined.index, aggregated_data_combined['Total Assets'], marker='o', linestyle='-', color='b')
plt.title('Total des actifs du secteur bancaire en Tunisie (2010-2032)')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux (en milliers de dinars)')
plt.grid(True)
plt.xticks(aggregated_data_combined.index, rotation=45)
plt.tight_layout()
plt.show()


# Création d'un DataFrame pour les prédictions
predictionsDA_df = pd.DataFrame({
    'Year': future_years.flatten(),
    'Predicted Total Assets': predictions
})

# Affichage du DataFrame
print(predictionsDA_df)

# Analyse des tendances
# Calcul du taux de croissance annuel des actifs
growth_rates = aggregated_data['Total Assets'].pct_change().dropna()
print("Taux de croissance annuel des actifs :", growth_rates)

# Identifier les années avec une forte croissance ou réduction
print("Années avec les taux de croissance les plus élevés :", growth_rates.nlargest(3))
print("Années avec les taux de croissance les plus faibles :", growth_rates.nsmallest(3))

# Affichage du DataFrame
# Création d'un DataFrame pour les prédictions
aggregated_data_combined = pd.DataFrame({
    'Year': future_years.flatten(),
    'Predicted Total Assets': predictions
})

# Affichage du DataFrame
print(predictionsDA_df)

# Création d'un DataFrame pour les années de 2010 à 2035 avec les valeurs prédites
years_all = np.arange(2010, 2036)
predictions_all = model.predict(years_all.reshape(-1, 1))

# Création d'un DataFrame avec les années et les valeurs prédites
all_data = pd.DataFrame({
    'Year': years_all,
    'Predicted Total Assets': predictions_all
})

# Affichage du DataFrame
print(all_data)
output_file_path = 'Alldatafusion.xlsx'
all_data.to_excel(output_file_path, index=False)
from sklearn.model_selection import train_test_split

# Partitionnement des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Entraînement du modèle
model.fit(X_train, y_train)

# Prédiction sur l'ensemble de test
y_pred = model.predict(X_test)

# Évaluation du modèle
mae = mean_absolute_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print("Erreur absolue moyenne (MAE) :", mae)
print("Coefficient de détermination (R2) :", r2)

from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline

# Préparation des données
X = all_data.index.values.reshape(-1, 1)  # Les années
y = all_data['Predicted Total Assets'].values  # Les actifs totaux

# Création d'une instance de modèle de régression polynomiale
degree = 3  # Vous pouvez ajuster le degré en fonction de la complexité des données
polyreg = make_pipeline(PolynomialFeatures(degree), LinearRegression())

# Entraînement du modèle
polyreg.fit(X, y)

# Prédiction sur l'ensemble des données (pour visualisation)
y_pred_poly = polyreg.predict(X)

# Affichage des résultats
plt.figure(figsize=(10, 6))
plt.scatter(X, y, color='blue')  # Données réelles
plt.plot(X, y_pred_poly, color='red')  # Prédiction du modèle
plt.title('Régression Polynomiale sur les Actifs Totaux du Secteur Bancaire')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux')
plt.show()

from statsmodels.tsa.arima.model import ARIMA
import statsmodels.api as sm

# Convertir l'index en série temporelle
ts = all_data['Predicted Total Assets']

# Définir et ajuster le modèle ARIMA
# p=2, d=1, q=2 sont des paramètres de départ, vous devrez peut-être les ajuster
model = ARIMA(ts, order=(2, 1, 2))
results = model.fit()

# Prévision
preds = results.forecast(steps=5)  # Prévoir les 5 prochaines années

# Affichage des prévisions
plt.figure(figsize=(10, 6))
plt.plot(ts.index, ts, label='Données Réelles')
plt.plot(np.arange(ts.index[-1] + 1, ts.index[-1] + 6), preds, label='Prévisions ARIMA', linestyle='--')
plt.title('Prévision des Actifs Totaux avec ARIMA')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux')
plt.legend()
plt.show()

import pmdarima as pm

# Chargement de la série temporelle
ts = all_data['Predicted Total Assets']

# Modélisation ARIMA avec auto_arima pour trouver automatiquement un bon ensemble de paramètres
model = pm.auto_arima(ts, start_p=1, start_q=1,
                      test='adf',       # Utilisation du test ADF pour trouver l'ordre de différenciation 'd'
                      max_p=5, max_q=5, # Plages maximales pour 'p' et 'q'
                      m=1,              # Fréquence de la série temporelle (m=1 pour des données annuelles)
                      d=None,           # Laisser auto_arima déterminer l'ordre de différenciation
                      seasonal=False,   # Pas de saisonnalité
                      start_P=0, 
                      D=0, 
                      trace=True,
                      error_action='ignore',  
                      suppress_warnings=True, 
                      stepwise=True)

print(model.summary())

# Prévision
forecast = model.predict(n_periods=5)

# Affichage des prévisions
future_years = np.arange(ts.index[-1] + 1, ts.index[-1] + 6)
plt.figure(figsize=(10, 6))
plt.plot(ts.index, ts, label='Données Historiques')
plt.plot(future_years, forecast, label='Prévisions ARIMA', color='red', linestyle='--')
plt.title('Prévisions des Actifs Totaux avec ARIMA Optimisé')
plt.xlabel('Année')
plt.ylabel('Actifs Totaux')
plt.legend()
plt.show()

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
banks = ['ATB', 'BNA', 'ATTIBK', 'BT', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB']  # Ajoutez plus de banques selon vos besoins

# Exemple de données (à remplacer par vos données réelles)
data = {
    'Year': np.repeat(np.arange(2010, 2035), len(banks)),
    'Bank': np.tile(banks, 2035-2010),
    'GDP': np.random.rand((2035-2010)*len(banks)),  # Exemple de données PIB
    'InterestRate': np.random.rand((2035-2010)*len(banks)),  # Exemple de taux d'intérêt
    'TotalAssets': np.random.rand((2035-2010)*len(banks)) * 10000  # Exemple d'actifs totaux
}

bank_data_df = pd.DataFrame(data)

# Encoder 'Bank' en variables numériques
bank_data_df = pd.get_dummies(bank_data_df, columns=['Bank'])
bank_data_df
# Séparation des caractéristiques et de la cible
X = bank_data_df.drop('TotalAssets', axis=1)
y = bank_data_df['TotalAssets']

# Division en ensembles d'entraînement et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Création et entraînement du modèle
model = LinearRegression()
model.fit(X_train, y_train)

from sklearn.metrics import mean_squared_error, mean_absolute_error

# Prédiction sur l'ensemble de test
y_pred = model.predict(X_test)
# Calcul du R²
r2 = r2_score(y_test, y_pred)
# Calcul de la MAE
mae = mean_absolute_error(y_test, y_pred)
# Calcul de la MSE
mse = mean_squared_error(y_test, y_pred)
print("Mean Squared Error (MSE) :", mse)
print("Mean Absolute Error (MAE) :", mae)
print("Coefficient de détermination R² :", r2)


import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score

# Chemin d'accès au fichier Excel - ajustez selon votre environnement
excel_path = 'EtudeSB - Copie.xlsx'
xls = pd.ExcelFile(excel_path)

# Liste des banques pour itérer
banks = ['ATB', 'BNA', 'ATTIBK', 'BT', 'AMENBK', 'BIAT', 'STB', 'UBCI', 'UIB']  # Exemple de banques

# Initialiser une liste pour stocker les performances des modèles
model_performance_list = []

# Itérer sur chaque banque pour entraîner un modèle et évaluer sa performance
for bank in banks:
    data_list = []
    for year in range(2010, 2023):  # De 2010 à 2022
        sheet_name = f'BILAN {year}'
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if bank in df.columns:
            total_assets = pd.to_numeric(df[bank], errors='coerce').fillna(0).sum()
            data_list.append({'Year': year, 'Total Assets': total_assets})
    
    # Créer un DataFrame pour la banque actuelle
    bank_data = pd.DataFrame(data_list)
    
    # Si la banque a des données disponibles
    if not bank_data.empty:
        X = bank_data[['Year']] - bank_data['Year'].min()  # Normalisation de l'année
        y = bank_data['Total Assets']
        
        # Division des données en ensemble d'entraînement et de test
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
        
        # Création et entraînement du modèle de régression linéaire
        model = LinearRegression()
        model.fit(X_train, y_train)
        
        # Prédiction sur l'ensemble de test et calcul du score R²
        y_pred = model.predict(X_test)
        r2 = r2_score(y_test, y_pred)
        
        # Stocker la performance du modèle dans la liste
        model_performance_list.append({'Bank': bank, 'R2_Score': r2})

# Convertir la liste en DataFrame
model_performance = pd.DataFrame(model_performance_list)

# Afficher les performances des modèles
print(model_performance)

import numpy as np
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt

# Supposons X_train, X_test, y_train, y_test sont déjà définis

models = {
    "Linear Regression": LinearRegression(),
    "KNN Regression": KNeighborsRegressor(),
    "Random Forest": RandomForestRegressor(),
    # Ajoutez ou remplacez par d'autres modèles ici
}

results = {}

for name, model in models.items():
    model.fit(X_train, y_train)  # Entraînement du modèle
    predictions = model.predict(X_test)  # Prédiction sur l'ensemble de test
    mse = mean_squared_error(y_test, predictions)  # Calcul du MSE
    r2 = r2_score(y_test, predictions)  # Calcul du score R2
    
    results[name] = {"MSE": mse, "R2": r2}  # Stockage des résultats

# Visualisation des résultats
labels, mse_scores, r2_scores = [], [], []
for model_name, metrics in results.items():
    labels.append(model_name)
    mse_scores.append(metrics["MSE"])
    r2_scores.append(metrics["R2"])

x = np.arange(len(labels))  # labels de localisation
width = 0.35  # largeur des barres

fig, ax = plt.subplots()
rects1 = ax.bar(x - width/2, mse_scores, width, label='MSE')
rects2 = ax.bar(x + width/2, r2_scores, width, label='R2')

# Ajout de textes pour labels, titre, axes ticks, etc.
ax.set_xlabel('Modèles')
ax.set_ylabel('Scores')
ax.set_title('Scores par modèle')
ax.set_xticks(x)
ax.set_xticklabels(labels)
ax.legend()

fig.tight_layout()

plt.show()
# Je vais d'abord importer les bibliothèques nécessaires pour travailler avec les données et les visualisations
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px

# Charger les données
file_path= 'EtudeSB - Copie.xlsx'
all_data = pd.read_excel(file_path, sheet_name=None)

# Initialiser un DataFrame pour le taux de croissance
credit_growth = pd.DataFrame()

# Calculer le taux de croissance du crédit pour chaque année et banque
for year in range(2011, 2023):
    current_year_data = all_data.get(f'BILAN {year}')
    previous_year_data = all_data.get(f'BILAN {year - 1}')
    
    if previous_year_data is not None:
        current_credits = current_year_data.set_index('En milliers de dinars').loc['Créances sur clientèle']
        previous_credits = previous_year_data.set_index('En milliers de dinars').loc['Créances sur clientèle']
        
        for bank in ['WIB', 'BARAKA']:
            if bank not in current_credits:
                current_credits[bank] = pd.NA
            if bank not in previous_credits:
                previous_credits[bank] = pd.NA
        
        growth_rate = (((current_credits / previous_credits) ** (1/1)) - 1) * 100
        credit_growth[year] = growth_rate

credit_growth.fillna(0, inplace=True)
credit_growth = credit_growth.round(2)
credit_growth = credit_growth.T

# Enregistrer les données calculées
output_file_path = 'Updated_EtudeSBF.xlsx'
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    credit_growth.to_excel(writer, sheet_name='Credit2')


## Visualisation 1: Taux de croissance du crédit
fig = px.line(credit_growth, labels={'value':'Taux de Croissance (%)', 'variable':'Banque', 'index':'Année'},
              title='Taux de Croissance Annuel du Crédit Bancaire par Banque')
fig.update_layout(xaxis_title='Année', yaxis_title='Taux de Croissance (%)', legend_title='Banque')


## Visualisation 2: Distribution des taux de croissance
for bank in credit_growth.columns:
    sns.histplot(credit_growth[bank], kde=True)
    plt.title(f'Distribution des taux de croissance pour {bank}')
    plt.xlabel('Taux de Croissance (%)')
    plt.ylabel('Fréquence')
    plt.show()


## Visualisation 3: Matrice de corrélation
correlation_matrix = credit_growth.corr()
sns.heatmap(correlation_matrix, annot=True)
plt.title('Matrice de corrélation des taux de croissance du crédit entre les banques')
plt.show()
## Visualisation 4: Boxplot de la distribution des taux de croissance
sns.boxplot(data=credit_growth)
plt.title("Distribution des Taux de Croissance du Crédit par Année")
plt.ylabel("Taux de Croissance (%)")
plt.xlabel("Année")
plt.xticks(rotation=45)
plt.show()
import pandas as pd

# Enregistrer les statistiques descriptives dans un fichier Excel
description_file = 'description_credit_growth_banque.xlsx'
with pd.ExcelWriter(description_file) as writer:
    print("Statistiques descriptives des taux de croissance du crédit :")
    print(credit_growth.describe())
    credit_growth.describe().to_excel(writer, sheet_name='Descriptives')

    # Classification de la croissance
    def classify_growth(rate):
        if rate < 5:
            return 'faible'
        elif rate < 15:
            return 'moyen'
        else:
            return 'fort'

    classified_growth = credit_growth.applymap(classify_growth)
    print(classified_growth)
    classified_growth.to_excel(writer, sheet_name='Classification')

print(f"Les statistiques descriptives et la classification ont été enregistrées dans '{description_file}'.")

# Visualisation de la classification
mapped_values = classified_growth.replace({'faible': 1, 'moyen': 2, 'fort': 3})
plt.figure(figsize=(12, 8))
sns.heatmap(mapped_values, annot=True, fmt=".0f", cmap='RdYlGn', cbar_kws={'ticks': [1, 2, 3]}, cbar=True)
plt.xlabel('Année')
plt.ylabel('Banque')
plt.title('Classification de la Croissance des Banques par Année')
cbar = plt.gca().collections[0].colorbar
cbar.set_ticklabels(['faible', 'moyen', 'fort'])
plt.show()
import pandas as pd

# Calcul de la moyenne de croissance et classification finale
mean_growth = credit_growth.mean(axis=0)
classification = mean_growth.apply(classify_growth)
final_table = pd.DataFrame({'Moyenne de Croissance': mean_growth, 'Classification': classification})

# Enregistrer la table finale dans un fichier Excel
final_table_file = 'final_table_credit_growth.xlsx'
final_table.to_excel(final_table_file, sheet_name='Final Table')

print(f"La table finale a été enregistrée dans '{final_table_file}'.")
print(final_table)
# Visualisation finale: Nombre de banques par catégorie de croissance
classification_count = classification.value_counts()
plt.figure(figsize=(10, 6))
sns.barplot(x=classification_count.index, y=classification_count.values, palette='RdYlGn')
plt.xlabel('Classification')
plt.ylabel('Nombre de Banques')
plt.title('Nombre de Banques par Catégorie de Croissance')
plt.show()
# Graphique à barres pour la moyenne de croissance et classification des banques
banques = final_table.index
moyenne_croissance = final_table['Moyenne de Croissance']
couleurs = final_table['Classification'].map({'faible': 'red', 'moyen': 'yellow', 'fort': 'green'})
plt.figure(figsize=(10, 8))
bars = plt.bar(banques, moyenne_croissance, color=couleurs)
plt.legend(handles=[plt.Rectangle((0,0),1,1, color=color) for color in ['red', 'yellow', 'green']],
           labels=['faible', 'moyen', 'fort'])
plt.title('Moyenne de Croissance et Classification des Banques')
plt.xlabel('Banque')
plt.ylabel('Moyenne de Croissance (%)')
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.show()

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# Charger les données
file_path = 'C:/Users/ASUS TUF I5/Downloads/Updated_EtudeSBF.xlsx'
credit_growth = pd.read_excel(file_path, sheet_name='Credit2', index_col=0)

# Classer la croissance selon des seuils prédéfinis
def classify_growth(growth_rate):
    if growth_rate < 5:
        return 'Faible'
    elif growth_rate < 15:
        return 'Moyenne'
    else:
        return 'Élevée'

classified_growth = credit_growth.applymap(classify_growth)

# Reformater les données pour la visualisation
classified_growth = classified_growth.stack().reset_index()
classified_growth.columns = ['Année', 'Banque', 'Classification']

# Visualiser la classification de la croissance du crédit
plt.figure(figsize=(10, 6))
sns.countplot(data=classified_growth, x='Année', hue='Classification')
plt.title('Classification de la Croissance du Crédit par Année')
plt.xlabel('Année')
plt.ylabel('Nombre de Banques')
plt.legend(title='Classification')
plt.show()


import pandas as pd

# Charger le fichier Excel
file_path= 'EtudeSB - Copie.xlsx'

# Lire toutes les feuilles du fichier Excel dans un dictionnaire
# Lire toutes les feuilles du fichier Excel dans un dictionnaire pour comprendre sa structure
all_data = pd.read_excel(file_path, sheet_name=None)

# Afficher les noms des feuilles pour comprendre la structure des données
sheet_names = list(all_data.keys())
sheet_names
# Initialiser un dictionnaire pour stocker le total des crédits par année
total_credits_per_year = {}

# Itérer sur chaque feuille de bilan pour calculer le total des crédits par année
for year in range(2010, 2023):
    sheet_name = f'BILAN {year}'
    
    # Vérifier si la feuille existe dans le fichier Excel
    if sheet_name in sheet_names:
        # Lire les données de la feuille de bilan de l'année courante
        current_year_data = all_data[sheet_name]
        
        # Vérifier si 'Créances sur clientèle' est dans les index après les avoir définis
        # Si oui, sommer toutes les valeurs pour obtenir le total des crédits de l'année
        if 'En milliers de dinars' in current_year_data.columns:
            current_year_data.set_index('En milliers de dinars', inplace=True)
            if 'Créances sur clientèle' in current_year_data.index:
                total_credits = current_year_data.loc['Créances sur clientèle'].sum()
                total_credits_per_year[year] = total_credits

# Convertir le dictionnaire en DataFrame pour un traitement ultérieur
total_credits_df = pd.DataFrame(list(total_credits_per_year.items()), columns=['Year', 'Total Credits'])

# Calculer le taux de croissance annuel du crédit
total_credits_df['Growth Rate'] = total_credits_df['Total Credits'].pct_change() * 100

total_credits_df
import matplotlib.pyplot as plt

# Créer un graphique pour visualiser la croissance du crédit bancaire par année
plt.figure(figsize=(12, 6))
plt.plot(total_credits_df['Year'], total_credits_df['Growth Rate'], marker='o', linestyle='-', color='b')
plt.title('Croissance Annuelle du Crédit Bancaire', fontsize=16)
plt.xlabel('Année', fontsize=14)
plt.ylabel('Taux de Croissance (%)', fontsize=14)
plt.grid(True, which='both', linestyle='--', linewidth=0.5)
plt.xticks(total_credits_df['Year'], rotation=45)
plt.tight_layout()

# Sauvegarder le graphique
#plt_path = '/mnt/data/credit_growth_annual.png'
#plt.savefig(plt_path)

plt.show() #plt_path
import pandas as pd

# Votre code existant pour calculer les taux de croissance annuels et créer un DataFrame
# (Assumant que total_credits_df contient les données que vous souhaitez enregistrer)

# Spécifiez le chemin où vous souhaitez enregistrer le fichier Excel
excel_path = 'croissance_credit_annuel.xlsx'

# Enregistrez les données dans un fichier Excel
total_credits_df.to_excel(excel_path, index=False)

# Affichez un message de confirmation
print(f"Les données ont été enregistrées dans '{excel_path}'.")


total_credits_df
import seaborn as sns

# Créer un histogramme pour visualiser la distribution des taux de croissance globaux
plt.figure(figsize=(10, 6))
sns.histplot(total_credits_df['Growth Rate'].dropna(), kde=True, color='skyblue')
plt.title('Distribution des Taux de Croissance Annuelle Globale du Crédit Bancaire', fontsize=14)
plt.xlabel('Taux de Croissance (%)', fontsize=12)
plt.ylabel('Fréquence', fontsize=12)
plt.grid(True, which='both', linestyle='--', linewidth=0.5)
plt.tight_layout()

# Sauvegarder le graphique

plt.show()

print(credit_growth)

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt

# Imaginons que `credit_growth` soit votre DataFrame existant contenant les taux de croissance du crédit par année jusqu'en 2022
# Voici un exemple de structure de ce DataFrame
data = {
    'Année': np.arange(2011, 2023),
    'Taux de Croissance (%)': np.random.rand(12) * 10  # Généré aléatoirement pour l'exemple
}
credit_growth = pd.DataFrame(data)

# Préparation des données pour le modèle de prédiction
X = credit_growth['Année'].values.reshape(-1, 1)  # Années
y = credit_growth['Taux de Croissance (%)'].values  # Taux de croissance

# Création et entraînement du modèle de régression linéaire
model = LinearRegression()
model.fit(X, y)

# Prédiction des taux de croissance de 2023 à 2035
future_years = np.arange(2023, 2036).reshape(-1, 1)
predicted_growth = model.predict(future_years)

# Création d'un nouveau DataFrame pour les prédictions
predictions_df = pd.DataFrame({
    'Année': np.arange(2023, 2036),
    'Taux de Croissance Prévu (%)': predicted_growth
})

# Affichage des prédictions
print(predictions_df)

# Enregistrement des prédictions dans un fichier Excel
# Note: Changez le chemin en fonction de votre système d'exploitation et de votre environnement
output_file_path = 'predictions_croissance_credit.xlsx'

predictions_df.to_excel(output_file_path, index=False)

print(f"Le fichier a été enregistré avec succès à l'emplacement suivant: {output_file_path}")

import pandas as pd

# Charger le fichier Excel
file_path= 'EtudeSB - Copie.xlsx'

# Lire toutes les feuilles du fichier Excel dans un dictionnaire
all_data = pd.read_excel(file_path, sheet_name=None)

# Initialiser un dictionnaire pour stocker le total des crédits par année
total_credits_per_year = {}

# Itérer sur chaque feuille de bilan pour calculer le total des crédits par année
for year in range(2010, 2023):
    sheet_name = f'BILAN {year}'
    
    # Vérifier si la feuille existe dans le fichier Excel
    if sheet_name in all_data:
        # Lire les données de la feuille de bilan de l'année courante
        current_year_data = all_data[sheet_name]
        
        # Vérifier si 'Créances sur clientèle' est dans les index après les avoir définis
        # Si oui, sommer toutes les valeurs pour obtenir le total des crédits de l'année
        if 'En milliers de dinars' in current_year_data.columns:
            current_year_data.set_index('En milliers de dinars', inplace=True)
            if 'Créances sur clientèle' in current_year_data.index:
                total_credits = current_year_data.loc['Créances sur clientèle'].sum()
                total_credits_per_year[year] = total_credits

# Convertir le dictionnaire en DataFrame pour un traitement ultérieur
total_credits_df = pd.DataFrame(list(total_credits_per_year.items()), columns=['Year', 'Total Credits'])

# Calculer le taux de croissance annuel du crédit
total_credits_df['Growth Rate'] = total_credits_df['Total Credits'].pct_change() * 100

# Classification de la croissance
def classify_growth(rate):
    if rate < 5:
        return 'faible'
    elif rate < 15:
        return 'moyen'
    else:
        return 'fort'

# Appliquer la classification au DataFrame
total_credits_df['Classification'] = total_credits_df['Growth Rate'].apply(classify_growth)

# Spécifier le chemin où vous souhaitez enregistrer le fichier Excel
excel_path = 'croissance_credit_annuel_année.xlsx'

# Enregistrer les données dans un fichier Excel
with pd.ExcelWriter(excel_path) as writer:
    # Enregistrer le DataFrame de taux de croissance
    total_credits_df.to_excel(writer, sheet_name='growth_rate', index=False)

    # Enregistrer la feuille de classification
    classification_sheet = total_credits_df[['Year', 'Classification']]
    classification_sheet.to_excel(writer, sheet_name='Classification', index=False)

# Afficher un message de confirmation
print(f"Les données ont été enregistrées dans '{excel_path}'.")

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression

# Supposons que `credit_growth` soit votre DataFrame existant contenant les taux de croissance du crédit par année jusqu'en 2022
# Voici un exemple de structure de ce DataFrame
data = {
    'Année': np.arange(2011, 2023),
    'Taux de Croissance (%)': np.random.rand(12) * 10  # Généré aléatoirement pour l'exemple
}
credit_growth = pd.DataFrame(data)

# Préparation des données pour le modèle de prédiction
X = credit_growth['Année'].values.reshape(-1, 1)  # Années
y = credit_growth['Taux de Croissance (%)'].values  # Taux de croissance

# Création et entraînement du modèle de régression linéaire
model = LinearRegression()
model.fit(X, y)

# Prédiction des taux de croissance de 2023 à 2035
future_years = np.arange(2023, 2036).reshape(-1, 1)
predicted_growth = model.predict(future_years)

# Création d'un nouveau DataFrame pour les prédictions
predictions_df = pd.DataFrame({
    'Année': np.arange(2023, 2036),
    'Taux de Croissance Prévu (%)': predicted_growth
})

# Classification de la croissance prédite
def classify_predicted_growth(rate):
    if rate < 5:
        return 'faible'
    elif rate < 15:
        return 'moyen'
    else:
        return 'fort'

predictions_df['Classification'] = predictions_df['Taux de Croissance Prévu (%)'].apply(classify_predicted_growth)

# Affichage des prédictions
print("Prévisions de taux de croissance par année :")
print(predictions_df)

# Enregistrement des prédictions dans un fichier Excel
output_file_path = 'predictions_croissance_credit.xlsx'
predictions_df.to_excel(output_file_path, index=False)

print(f"Le fichier a été enregistré avec succès à l'emplacement suivant: {output_file_path}")

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt

# Supposons que credit_growth est votre DataFrame contenant la croissance du crédit pour chaque banque et chaque année
# Assurez-vous que le DataFrame est structuré de sorte que chaque colonne représente une banque et chaque ligne une année

# Création d'un modèle de prédiction pour chaque banque
predictions_df_list = []
for bank in credit_growth.columns:
    # Préparation des données pour le modèle de prédiction
    X = np.arange(2011, 2023).reshape(-1, 1)  # Les années
    y = credit_growth[bank].values  # Taux de croissance pour la banque courante
    
    # Création et entraînement du modèle de régression linéaire
    model = LinearRegression()
    model.fit(X, y)
    
    # Prédiction des taux de croissance de 2023 à 2035 pour la banque courante
    future_years = np.arange(2023, 2036).reshape(-1, 1)
    predicted_growth = model.predict(future_years)
    
    # Création d'un DataFrame pour les prédictions de la banque courante
    bank_predictions_df = pd.DataFrame({
        'Année': np.arange(2023, 2036),
        f'Taux de Croissance Prévu (%) - {bank}': predicted_growth
    })
    
    predictions_df_list.append(bank_predictions_df)

# Concaténation de tous les DataFrames de prédictions dans un seul DataFrame
final_predictions_df = pd.concat(predictions_df_list, axis=1)

# Affichage des prédictions
print(final_predictions_df)

# Enregistrement des prédictions dans un fichier Excel
output_file_path = 'predictions_croissance_credit_par_banques.xlsx'
final_predictions_df.to_excel(output_file_path, index=False)
print(f"Le fichier a été enregistré avec succès à l'emplacement suivant: {output_file_path}")

import pandas as pd

# Charger le fichier Excel
file_path = 'Credit.xlsx'
credit_growth_df = pd.read_excel(file_path)

# Afficher les premières lignes pour vérifier la structure
credit_growth_df.head()

# Initialisation du modèle de régression linéaire
model = LinearRegression()

# Dictionnaire pour stocker les prédictions de chaque banque
predictions = {}

# Boucle sur chaque colonne (banque) du DataFrame, sauf la première colonne des années
for bank in credit_growth_df.columns[1:]:
    # Préparation des données
    X = credit_growth_df['Unnamed: 0'].values.reshape(-1, 1)  # Les années
    y = credit_growth_df[bank].values  # Taux de croissance du crédit pour la banque courante
    
    # Entraînement du modèle
    model.fit(X, y)
    
    # Prédiction des taux de croissance du crédit de 2023 à 2035 pour la banque courante
    predicted_growth_rates = model.predict(future_years)
    
    # Stockage des prédictions
    predictions[bank] = predicted_growth_rates

# Conversion du dictionnaire de prédictions en DataFrame
predictions_df = pd.DataFrame(predictions, index=np.arange(2023, 2036))

predictions_df

# Chemin vers le fichier de sortie
file_path = 'CreditBanque.xlsx'

# Enregistrer le DataFrame dans un fichier Excel
predictions_df.to_excel(file_path, index=True)
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsRegressor
from sklearn.metrics import mean_squared_error, r2_score

# Chemin du fichier Excel
excel_path = 'changee.xlsx'

# Nom de la feuille de calcul à lire
sheet_name = 'Feuil1'

# Lecture du fichier Excel dans un DataFrame
change_df = pd.read_excel(excel_path, sheet_name=sheet_name)

# Filtration des colonnes nécessaires
columns_to_keep = ['Dollar des USA', 'Yen Japonais', 'EURO']
date_column = [col for col in change_df.columns if 'Date' in col][0]
change_df_filtered = change_df[[date_column] + columns_to_keep]

# Nettoyage des données
change_df_filtered[date_column] = pd.to_datetime(change_df_filtered[date_column], errors='coerce')
change_df_filtered = change_df_filtered.dropna().replace(',', '.', regex=True)

for col in columns_to_keep:
    change_df_filtered[col] = change_df_filtered[col].astype(float)

# Affichage des premières lignes pour vérification
print(change_df_filtered.head())


# Sélection de la variable à prédire
target = 'Dollar des USA'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['Yen Japonais', 'EURO']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")


# Sélection de la variable à prédire
target = 'EURO'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['Yen Japonais', 'Dollar des USA']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")

# Sélection de la variable à prédire
target = 'Yen Japonais'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['EURO', 'Dollar des USA']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")

import pandas as pd
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsRegressor
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.pipeline import Pipeline
from sklearn.feature_selection import SelectKBest, f_regression

# Chargement des données
excel_path = 'changee.xlsx'
sheet_name = 'Feuil1'
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# Sélection et nettoyage des données (supposons que cette étape est déjà bien définie)
# Assuming 'Yen Japonais' and 'EURO' are the columns that need correction
# Example of converting columns to numeric if not already
df['Yen Japonais'] = pd.to_numeric(df['Yen Japonais'], errors='coerce')
df['EURO'] = pd.to_numeric(df['EURO'], errors='coerce')


# Division des données en ensembles d'apprentissage et de test
X = df[['Yen Japonais', 'EURO']]  # Supposons ces comme caractéristiques
y = df['Dollar des USA']  # Variable cible
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Construction du pipeline
pipeline = Pipeline([
    ('scaler', StandardScaler()),
    ('feature_selection', SelectKBest(score_func=f_regression, k='all')),
    ('knn', KNeighborsRegressor())
])

# Sélection du paramètre K via validation croisée
param_grid = {'knn__n_neighbors': range(1, 30)}
grid_search = GridSearchCV(pipeline, param_grid, cv=5, scoring='neg_mean_squared_error')
grid_search.fit(X_train, y_train)

# Meilleur modèle et paramètres
print("Meilleur paramètre K:", grid_search.best_params_)
# Check for NaN values
print(df.isnull().sum())

# Example way to handle NaN values, such as filling with the mean of the column
df = df.fillna(df.mean())

# Évaluation du modèle
y_pred = grid_search.predict(X_test)
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")
import pandas as pd

# Chemin du fichier Excel
excel_path = r"C:\Users\ASUS TUF I5\Desktop\PFE\ETUDE DU SECTEUR\changee.xlsx"

# Nom de la feuille de calcul à lire
sheet_name = 'Feuil1'

# Lecture du fichier Excel dans un DataFrame
change_df_cleaned = pd.read_excel(excel_path, sheet_name=sheet_name)

# Liste des colonnes à conserver
columns_to_keep = ['Dollar des USA', 'Yen Japonais', 'EURO']

# Vérifier si les noms de colonne existent dans le DataFrame
if all(col in change_df_cleaned.columns for col in columns_to_keep):
    # Filtrer le DataFrame
    change_df_filtered = change_df_cleaned[columns_to_keep]
else:
    # Afficher un message d'erreur
    print("Les noms de colonne ne correspondent pas. Veuillez vérifier.")

# Le nom de la colonne de date est différent et doit être inclus également pour l'analyse temporelle
# Nous allons trouver le nom de la colonne contenant les informations de date
date_column = change_df_cleaned.columns[change_df_cleaned.columns.str.contains('Date')][0]

# Inclure la colonne de date dans notre DataFrame filtré
change_df_filtered_with_date = change_df_cleaned[[date_column] + columns_to_keep]

# Nettoyage des données :
# - Conversion de la colonne de date en datetime
# - Remplacement des virgules par des points dans les valeurs de taux de change pour les convertir en type float
change_df_filtered_with_date[date_column] = pd.to_datetime(change_df_filtered_with_date[date_column], errors='coerce')
change_df_filtered_with_date = change_df_filtered_with_date.replace(',', '.', regex=True).dropna()

# Conversion des valeurs de taux de change de string en float
for col in columns_to_keep:
    try:
        change_df_filtered_with_date[col] = change_df_filtered_with_date[col].astype(float)
    except ValueError as e:
        print(f"Erreur lors de la conversion de la colonne {col}: {e}")

# Affichage des premières lignes du DataFrame nettoyé
change_df_filtered_with_date.head()

import pandas as pd
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import adfuller, acf, pacf
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.arima.model import ARIMA

# Remove commas and convert values to float
change_df_filtered = change_df_filtered.replace(',', '.', regex=True)
change_df_filtered = change_df_filtered.astype(float)

# Continue with your analysis as before
for currency in columns_to_keep:
    print(f'\nAnalyse de {currency}:')
    result = adfuller(change_df_filtered[currency])
    print(f'Statistique ADF : {result[0]}')
    print(f'P-value : {result[1]}')
    print('Valeurs Critiques :')
    for key, value in result[4].items():
        print(f'    {key}: {value}')


import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import adfuller, acf, pacf
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.arima.model import ARIMA
# Remplacement des virgules par des points et conversion des chaînes de caractères en nombres flottants

for currency in columns_to_keep:
    
    print(f'\nAnalyse de {currency}:')
    result = adfuller(change_df_filtered[currency])
    print(f'Statistique ADF : {result[0]}')
    print(f'P-value : {result[1]}')
    print('Valeurs Critiques :')
    for key, value in result[4].items():
        print(f'    {key}: {value}')



for currency in columns_to_keep:
    # Préparation de la série différenciée si nécessaire
    if result[1] > 0.05:
        change_df_filtered[f'{currency}_diff'] = change_df_filtered[currency].diff().dropna()
        diff_column = f'{currency}_diff'
    else:
        diff_column = currency

    # Calculer un nombre approprié de lags : min entre 10 et la moitié de la taille de l'échantillon
    max_lags = min(10, len(change_df_filtered[diff_column].dropna()) // 2 - 1)

    # Vérifier si max_lags est suffisamment grand
    if max_lags > 0:
        # Tracer ACF et PACF avec le nombre ajusté de lags
        plt.figure(figsize=(14, 7))
        plt.subplot(121)
        plot_acf(change_df_filtered[diff_column].dropna(), ax=plt.gca(), lags=max_lags)
        plt.title(f'ACF de {currency}')
        plt.subplot(122)
        plot_pacf(change_df_filtered[diff_column].dropna(), ax=plt.gca(), lags=max_lags)
        plt.title(f'PACF de {currency}')
        plt.tight_layout()
        plt.show()
    else:
        print(f'Pas assez de données pour tracer ACF et PACF pour {currency}')

for currency in columns_to_keep:
    # Construction et ajustement du modèle ARIMA
    if 'diff' in diff_column:  # Si une différenciation a été effectuée
        model_arima = ARIMA(change_df_filtered[currency], order=(1, 1, 1))
    else:
        model_arima = ARIMA(change_df_filtered[currency], order=(1, 0, 1))
    model_arima_fit = model_arima.fit()

    # Diagnostic du modèle
    model_arima_fit.plot_diagnostics(figsize=(14, 8))
    plt.show()

    # Affichage du résumé du modèle
    print(model_arima_fit.summary())

from statsmodels.tsa.arima.model import ARIMA
import itertools
import warnings

warnings.filterwarnings("ignore")  # Ignorer les avertissements

# Devise à analyser
currencies = ['Dollar des USA', 'EURO', 'Yen Japonais']

# Définition de la grille de recherche
p = d = q = range(0, 3)  # Plage de paramètres
pdq_combinations = list(itertools.product(p, d, q))

best_params = {}

for currency in currencies:
    best_aic = float("inf")
    best_combination = (0, 0, 0)
    for combination in pdq_combinations:
        try:
            model = ARIMA(change_df_filtered_with_date[currency], order=combination)
            model_fit = model.fit()
            if model_fit.aic < best_aic:
                best_aic = model_fit.aic
                best_combination = combination
        except:
            continue
    best_params[currency] = best_combination
    print(f"{currency} - Meilleure combinaison : {best_combination} avec un AIC de {best_aic}")

from statsmodels.tsa.arima.model import ARIMA

# Ajustement du modèle ARIMA avec les paramètres optimisés
model_optimized = ARIMA(change_df_filtered_with_date['Dollar des USA'], order=(0, 2, 1))
model_optimized_fit = model_optimized.fit()

# Affichage du résumé du modèle pour inspection
print(model_optimized_fit.summary())
from sklearn.metrics import r2_score

# Prévisions sur les données historiques
predictions = model_optimized_fit.predict(start=change_df_filtered_with_date.index[0], end=change_df_filtered_with_date.index[-1], dynamic=False)

# Calcul de R^2
r2 = r2_score(change_df_filtered_with_date['Dollar des USA'][2:], predictions[2:])  # On exclut les premiers points à cause de la différenciation d'ordre 2
print(f"R^2 score: {r2:.4f}")

from statsmodels.tsa.arima.model import ARIMA

# Ajustement du modèle ARIMA avec les paramètres optimisés
model_optimized = ARIMA(change_df_filtered_with_date['EURO'], order=(1, 1, 0))
model_optimized_fit = model_optimized.fit()

# Affichage du résumé du modèle pour inspection
print(model_optimized_fit.summary())
from sklearn.metrics import r2_score

# Prévisions sur les données historiques
predictions = model_optimized_fit.predict(start=change_df_filtered_with_date.index[0], end=change_df_filtered_with_date.index[-1], dynamic=False)

# Calcul de R^2
r2 = r2_score(change_df_filtered_with_date['EURO'][2:], predictions[2:])  # On exclut les premiers points à cause de la différenciation d'ordre 2
print(f"R^2 score: {r2:.4f}")


from statsmodels.tsa.arima.model import ARIMA

# Ajustement du modèle ARIMA avec les paramètres optimisés
model_optimized = ARIMA(change_df_filtered_with_date['Yen Japonais'], order=(0,2,0))
model_optimized_fit = model_optimized.fit()

# Affichage du résumé du modèle pour inspection
print(model_optimized_fit.summary())
from sklearn.metrics import r2_score

# Prévisions sur les données historiques
predictions = model_optimized_fit.predict(start=change_df_filtered_with_date.index[0], end=change_df_filtered_with_date.index[-1], dynamic=False)

# Calcul de R^2
r2 = r2_score(change_df_filtered_with_date['Yen Japonais'][2:], predictions[2:])  # On exclut les premiers points à cause de la différenciation d'ordre 2
print(f"R^2 score: {r2:.4f}")


import pandas as pd
import matplotlib.pyplot as plt
from statsmodels.tsa.arima.model import ARIMA
import numpy as np

# Je suppose que votre DataFrame est déjà nettoyé et contient les colonnes nécessaires
# et que vous avez déjà défini `change_df_filtered_with_date` avec l'index correct.

# Définir l'index de `change_df_filtered_with_date` sur la colonne de date si ce n'est pas déjà fait
change_df_filtered_with_date.set_index(date_column, inplace=True)

# Assurez-vous que l'index est de type datetime
change_df_filtered_with_date.index = pd.to_datetime(change_df_filtered_with_date.index)

# Le reste de votre code pour l'ARIMA et les prévisions reste le même
optimal_params = {
    'Dollar des USA': (0, 2, 1),
    'EURO': (1, 1, 0),
    'Yen Japonais': (0, 2, 0)
}

# Forecast steps
forecast_steps = 5
# Placeholder pour les valeurs prévues
forecasts = {}

# Prévision pour chaque devise
for currency, params in optimal_params.items():
    model = ARIMA(change_df_filtered_with_date[currency], order=params)
    model_fit = model.fit()
    forecasts[currency] = model_fit.forecast(steps=forecast_steps)

# Création d'un DataFrame à partir des prévisions
forecast_dates = pd.date_range(start=change_df_filtered_with_date.index[-1] + pd.DateOffset(days=1), 
                               periods=forecast_steps, freq='Y')
forecasts_df = pd.DataFrame(forecasts, index=forecast_dates)

# Tracé
fig, axs = plt.subplots(len(optimal_params), 1, figsize=(10, 15))
for i, (currency, forecast) in enumerate(forecasts.items()):
    axs[i].plot(change_df_filtered_with_date.index, change_df_filtered_with_date[currency], label='Historical')
    axs[i].plot(forecast_dates, forecast, label='Forecast', linestyle='--')
    axs[i].set_title(f'{currency} Forecast')
    axs[i].set_xlabel('Year')
    axs[i].set_ylabel('Exchange Rate')
    axs[i].legend()

plt.tight_layout()
# Afficher le DataFrame des prévisions
forecasts_df


from sklearn.metrics import mean_squared_error
from math import sqrt
import numpy as np

# Génération de données fictives pour les devises
np.random.seed(42)  # Pour la reproductibilité
data_length = 100  # Total de points de données
train_length = 80  # Points de données pour l'entraînement

# Génération de séries temporelles aléatoires pour chaque devise
data_dollar = np.random.randn(data_length).cumsum()
data_euro = np.random.randn(data_length).cumsum()
data_yen = np.random.randn(data_length).cumsum()

# Division en ensembles d'entraînement et de test
train_dollar, test_dollar = data_dollar[:train_length], data_dollar[train_length:]
train_euro, test_euro = data_euro[:train_length], data_euro[train_length:]
train_yen, test_yen = data_yen[:train_length], data_yen[train_length:]

# Dictionnaire pour stocker les erreurs MSE pour chaque devise
mse_errors = {}

# Fonction pour ajuster un modèle ARIMA et calculer le MSE
def evaluate_arima(data_train, data_test, order=(1,1,1)):
    model = ARIMA(data_train, order=order)
    model_fit = model.fit()
    predictions = model_fit.forecast(steps=len(data_test))
    mse = mean_squared_error(data_test, predictions)
    return mse

# Évaluation pour chaque devise
mse_errors['Dollar des USA'] = evaluate_arima(train_dollar, test_dollar)
mse_errors['EURO'] = evaluate_arima(train_euro, test_euro)
mse_errors['Yen Japonais'] = evaluate_arima(train_yen, test_yen)
# Dictionnaire pour stocker les scores R^2 pour chaque devise


mse_errors

from sklearn.metrics import r2_score
from statsmodels.tsa.arima.model import ARIMA

# Fonction pour ajuster un modèle ARIMA et calculer le R^2
def evaluate_arima_r2(data_train, data_test, order=(1,1,1)):
    model = ARIMA(data_train, order=order)
    model_fit = model.fit()
    predictions = model_fit.forecast(steps=len(data_test))
    r2 = r2_score(data_test, predictions)
    return r2

# Initialisation du dictionnaire pour stocker les scores R^2
r2_scores = {}

# Évaluation R^2 pour chaque devise
r2_scores['Dollar des USA'] = evaluate_arima_r2(train_dollar, test_dollar)
r2_scores['EURO'] = evaluate_arima_r2(train_euro, test_euro)
r2_scores['Yen Japonais'] = evaluate_arima_r2(train_yen, test_yen)

print(r2_scores)

import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt


# Paramètres pour la simulation
np.random.seed(42)  # Pour la reproductibilité
n_samples = 100
slope = 3
intercept = 50
time = np.arange(n_samples).reshape(-1, 1)

# Simuler des données pour chaque devise
currencies = ['Dollar des USA', 'EURO', 'Yen Japonais']
data = {}
for currency in currencies:
    noise = np.random.normal(0, 5, n_samples)  # Bruit ajouté
    data[currency] = slope * time.flatten() + intercept + noise

# Conversion en DataFrame
df = pd.DataFrame(data, columns=currencies)

from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split

r2_scores = {}
forecasts_df = {}

for currency in currencies:
    # Préparation des données
    X = time  # variable indépendante
    y = df[currency].values.reshape(-1, 1)  # variable dépendante
    
    # Normalisation des données
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    y_scaled = scaler.fit_transform(y)
    
    # Séparation en données d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X_scaled, y_scaled, test_size=0.2, random_state=42)
    
    # Création et ajustement du modèle
    model = LinearRegression()
    model.fit(X_train, y_train)
    
    # Prédiction sur les données de test
    y_pred = model.predict(X_test)
    r2_scores[currency] = r2_score(y_test, y_pred)
    
    # Stockage des prédictions
    forecasts_df[currency] = model.predict(X_scaled).flatten()

# Affichage des R²
print("R² Scores:", r2_scores)

# Conversion des prédictions en DataFrame pour affichage
forecasts_df = pd.DataFrame(forecasts_df, columns=currencies)

from sklearn.metrics import mean_squared_error

# Stockage des MSE
mse_scores = {}

for currency in currencies:
    # Calcul de la MSE
    mse = mean_squared_error(y_scaled, model.predict(X_scaled))
    mse_scores[currency] = mse

# Affichage des MSE
print("MSE Scores:", mse_scores)

import pandas as pd

# Chemin du fichier Excel
excel_path = r"C:\Users\ASUS TUF I5\Desktop\PFE\ETUDE DU SECTEUR\changee.xlsx"

# Nom de la feuille de calcul à lire
sheet_name = 'Feuil1'

# Lecture du fichier Excel dans un DataFrame
change_df_cleaned = pd.read_excel(excel_path, sheet_name=sheet_name)

# Liste des colonnes à conserver
columns_to_keep = ['Dollar des USA', 'Yen Japonais', 'EURO']

# Vérifier si les noms de colonne existent dans le DataFrame
if all(col in change_df_cleaned.columns for col in columns_to_keep):
    # Filtrer le DataFrame
    change_df_filtered = change_df_cleaned[columns_to_keep]
else:
    # Afficher un message d'erreur
    print("Les noms de colonne ne correspondent pas. Veuillez vérifier.")

# Le nom de la colonne de date est différent et doit être inclus également pour l'analyse temporelle
# Nous allons trouver le nom de la colonne contenant les informations de date
date_column = change_df_cleaned.columns[change_df_cleaned.columns.str.contains('Date')][0]

# Inclure la colonne de date dans notre DataFrame filtré
change_df_filtered_with_date = change_df_cleaned[[date_column] + columns_to_keep]

# Nettoyage des données :
# - Conversion de la colonne de date en datetime
# - Remplacement des virgules par des points dans les valeurs de taux de change pour les convertir en type float
change_df_filtered_with_date[date_column] = pd.to_datetime(change_df_filtered_with_date[date_column], errors='coerce')
change_df_filtered_with_date = change_df_filtered_with_date.replace(',', '.', regex=True).dropna()

# Conversion des valeurs de taux de change de string en float
for col in columns_to_keep:
    try:
        change_df_filtered_with_date[col] = change_df_filtered_with_date[col].astype(float)
    except ValueError as e:
        print(f"Erreur lors de la conversion de la colonne {col}: {e}")

# Affichage des premières lignes du DataFrame nettoyé
change_df_filtered_with_date.head()

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from datetime import datetime

# Assurez-vous d'avoir le DataFrame 'change_df_filtered_with_date' prêt comme spécifié dans votre code initial

# Conversion de la colonne de date en une valeur numérique (nombre de jours depuis une date)
change_df_filtered_with_date['date_numeric'] = (change_df_filtered_with_date[date_column] - pd.Timestamp("1970-01-01")).dt.days

results = [] # Pour stocker les résultats

for currency in columns_to_keep:
    X = change_df_filtered_with_date[['date_numeric']]  # Features
    y = change_df_filtered_with_date[currency]  # Cible

    # Division des données en ensembles d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Construction et entraînement du modèle de régression linéaire
    model = LinearRegression()
    model.fit(X_train, y_train)

    # Prédiction sur l'ensemble de test
    y_pred = model.predict(X_test)

    # Calcul des métriques d'évaluation
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))
    r2 = r2_score(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)

    # Stockage des résultats
    results.append({
        'Currency': currency,
        'RMSE': rmse,
        'R2': r2,
        'MAE': mae,
        'Predictions': y_pred,
        'Actual': y_test.values,
        'Test Dates': X_test['date_numeric'].apply(lambda x: pd.Timestamp("1970-01-01") + pd.Timedelta(days=x)).dt.year
    })

# Affichage des résultats pour chaque devise
for result in results:
    print(f"\nCurrency: {result['Currency']}")
    df_results = pd.DataFrame({
        'Year': result['Test Dates'],
        'Actual': result['Actual'],
        'Predicted': result['Predictions']
    })
    print(df_results.head())  # Affichage des premières lignes pour l'exemple
    print(f"RMSE: {result['RMSE']}, R2: {result['R2']}, MAE: {result['MAE']}")


# Assumons que 'change_df_filtered_with_date' et 'columns_to_keep' sont déjà préparés
# Assurons-nous également que la colonne des dates est au format datetime
change_df_filtered_with_date[date_column] = pd.to_datetime(change_df_filtered_with_date[date_column])

# Ajout d'une colonne pour l'année
change_df_filtered_with_date['Year'] = change_df_filtered_with_date[date_column].dt.year

from sklearn.linear_model import LinearRegression
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
# Initialisation d'un DataFrame pour stocker les prédictions futures
future_years = np.arange(2024, 2036)
predictions = pd.DataFrame(future_years, columns=['Year'])

# Boucle sur chaque devise pour entraîner un modèle et faire des prédictions
for currency in columns_to_keep:
    # Sélection des données pour la devise actuelle
    X = change_df_filtered_with_date[['Year']]
    y = change_df_filtered_with_date[currency]

    # Conversion de l'année en format numérique pour l'entraînement
    X = X.values.reshape(-1, 1)  # Convertir en format attendu par sklearn
    
    # Entraînement du modèle de régression linéaire
    model = LinearRegression()
    model.fit(X, y)

    # Préparation des données pour les futures prédictions (2023 à 2028)
    future_X = future_years.reshape(-1, 1)

    # Prédiction pour les années futures
    future_preds = model.predict(future_X)
    predictions[currency] = future_preds

# Affichage des prédictions pour chaque devise
print(predictions)
predictions.to_excel("predictions_devises.xlsx", index=False)

# Save the predictions to an Excel file
predictions_output_path = 'future_predictions.xlsx'
predictions.to_excel(predictions_output_path, index=False)

# Tracé des prédictions futures pour chaque devise
for currency in columns_to_keep:
    plt.plot(predictions['Year'], predictions[currency], label=currency)

plt.title('Prédictions des taux de change de 2023 à 2028')
plt.xlabel('Année')
plt.ylabel('Taux de change')
plt.legend()
plt.show()

from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import numpy as np

# Initialiser un dictionnaire pour stocker les métriques pour chaque devise
metrics = {}

for currency in columns_to_keep:
    # Préparation des données
    X = change_df_filtered_with_date['Year'].values.reshape(-1, 1)
    y = change_df_filtered_with_date[currency].values
    
    # Division des données en ensembles d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Construction et entraînement du modèle de régression linéaire
    model = LinearRegression()
    model.fit(X_train, y_train)

    # Prédiction sur l'ensemble de test
    y_pred = model.predict(X_test)

    # Calcul des métriques
    r2 = r2_score(y_test, y_pred)
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)

    # Stockage des métriques
    metrics[currency] = {'R2': r2, 'MSE': mse, 'MAE': mae}

# Affichage des métriques pour chaque devise
for currency, metric_values in metrics.items():
    print(f"Devise: {currency}")
    for metric, value in metric_values.items():
        print(f"{metric}: {value}")
    print()  # Ligne vide pour la séparation

import matplotlib.pyplot as plt

# Assurez-vous que 'columns_to_keep' et 'change_df_filtered_with_date' sont définis comme avant

for currency in columns_to_keep:
    # Préparation des données
    X = change_df_filtered_with_date['Year'].values.reshape(-1, 1)
    y = change_df_filtered_with_date[currency].values
    
    # Division des données
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Entraînement du modèle
    model = LinearRegression()
    model.fit(X_train, y_train)

    # Prédiction sur l'ensemble de test
    y_pred = model.predict(X_test)

    # Tracé des données réelles vs prédictions
    plt.figure(figsize=(10, 6))
    plt.scatter(X_test, y_test, color='blue', label='Données réelles', alpha=0.6)
    plt.scatter(X_test, y_pred, color='red', label='Prédictions', alpha=0.6)
    plt.title(f"Données réelles vs Prédictions pour {currency}")
    plt.xlabel('Année')
    plt.ylabel(f'Taux de change pour {currency}')
    plt.legend()
    plt.show()

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsRegressor
from sklearn.metrics import mean_squared_error, r2_score

# Chemin du fichier Excel
excel_path = 'changee.xlsx'

# Nom de la feuille de calcul à lire
sheet_name = 'Feuil1'

# Lecture du fichier Excel dans un DataFrame
change_df = pd.read_excel(excel_path, sheet_name=sheet_name)

# Filtration des colonnes nécessaires
columns_to_keep = ['Dollar des USA', 'Yen Japonais', 'EURO']
date_column = [col for col in change_df.columns if 'Date' in col][0]
change_df_filtered = change_df[[date_column] + columns_to_keep]

# Nettoyage des données
change_df_filtered[date_column] = pd.to_datetime(change_df_filtered[date_column], errors='coerce')
change_df_filtered = change_df_filtered.dropna().replace(',', '.', regex=True)

for col in columns_to_keep:
    change_df_filtered[col] = change_df_filtered[col].astype(float)

# Affichage des premières lignes pour vérification
print(change_df_filtered.head())

# Sélection de la variable à prédire
target = 'Dollar des USA'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['Yen Japonais', 'EURO']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

from sklearn.neighbors import KNeighborsRegressor

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")

# Sélection de la variable à prédire
target = 'EURO'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['Yen Japonais', 'Dollar des USA']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")

# Sélection de la variable à prédire
target = 'Yen Japonais'

# Séparation des caractéristiques et de la cible
X = change_df_filtered[['EURO', 'Dollar des USA']]
y = change_df_filtered[target]

# Division des données
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Normalisation
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Création du modèle KNN
knn = KNeighborsRegressor(n_neighbors=5)

# Entraînement du modèle
knn.fit(X_train_scaled, y_train)

# Prédictions
y_pred = knn.predict(X_test_scaled)

# Calcul de l'erreur quadratique moyenne et du coefficient de détermination
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

print(f"MSE: {mse}")
print(f"R^2: {r2}")

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from statsmodels.tsa.arima.model import ARIMA
from sklearn.neighbors import KNeighborsRegressor

# Assurez-vous d'avoir le DataFrame 'change_df_filtered_with_date' prêt comme spécifié dans votre code initial

# Définition des colonnes à conserver pour l'analyse
columns_to_keep = ['Dollar des USA', 'EURO', 'Yen Japonais']

# Initialisation d'un dictionnaire pour stocker les métriques pour chaque devise et méthode
metrics = {}

# Régression linéaire
for currency in columns_to_keep:
    X = change_df_filtered_with_date[['Year']].values
    y = change_df_filtered_with_date[currency].values
    
    # Division des données en ensembles d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Construction et entraînement du modèle de régression linéaire
    model = LinearRegression()
    model.fit(X_train, y_train)

    # Prédiction sur l'ensemble de test
    y_pred = model.predict(X_test)

    # Calcul des métriques
    r2 = r2_score(y_test, y_pred)
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)

    # Stockage des métriques
    metrics[f'{currency} - Linear Regression'] = {'R2': r2, 'MSE': mse, 'MAE': mae}

# ARIMA
for currency in columns_to_keep:
    # Ajustement du modèle ARIMA avec les paramètres optimisés
    model_optimized = ARIMA(change_df_filtered_with_date[currency], order=(1, 1, 0))
    model_optimized_fit = model_optimized.fit()

    # Prévisions sur les données historiques
    predictions = model_optimized_fit.predict(start=change_df_filtered_with_date.index[0],
                                               end=change_df_filtered_with_date.index[-1], dynamic=False)

    # Calcul de R^2
    r2 = r2_score(change_df_filtered_with_date[currency][1:], predictions[1:])  # On exclut les premiers points à cause de la différenciation d'ordre 1

    # Stockage des métriques
    metrics[f'{currency} - ARIMA'] = {'R2': r2}

# KNN
for currency in columns_to_keep:
    # Sélection des données pour la devise actuelle
    X = change_df_filtered_with_date[['Year']].values
    y = change_df_filtered_with_date[currency].values

    # Division des données en ensembles d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Création et ajustement du modèle KNN
    knn = KNeighborsRegressor(n_neighbors=5)
    knn.fit(X_train, y_train)

    # Prédiction sur l'ensemble de test
    y_pred = knn.predict(X_test)

    # Calcul des métriques
    r2 = r2_score(y_test, y_pred)
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)

    # Stockage des métriques
    metrics[f'{currency} - KNN'] = {'R2': r2, 'MSE': mse, 'MAE': mae}

# Affichage des métriques pour chaque devise et méthode
for method, metric_values in metrics.items():
    print(f"Method: {method}")
    for metric, value in metric_values.items():
        print(f"{metric}: {value}")
    print()  # Ligne vide pour la séparation

import matplotlib.pyplot as plt

# Initialisation des listes pour stocker les métriques
methods = []
r2_scores = []
mse_scores = []
mae_scores = []

# Parcourir les métriques et les stocker dans les listes
for method, metric_values in metrics.items():
    methods.append(method)
    r2_scores.append(metric_values['R2'])
    mse_scores.append(metric_values.get('MSE', 0))  # Si le score MSE n'est pas disponible, on utilise 0
    mae_scores.append(metric_values.get('MAE', 0))  # Si le score MAE n'est pas disponible, on utilise 0

# Création du graphique comparatif
plt.figure(figsize=(10, 6))

# R² Scores
plt.barh(methods, r2_scores, color='skyblue', label='R²')

# MSE Scores
if any(mse_scores):  # Vérifie s'il y a des scores MSE disponibles
    plt.barh(methods, mse_scores, color='lightgreen', label='MSE')

# MAE Scores
if any(mae_scores):  # Vérifie s'il y a des scores MAE disponibles
    plt.barh(methods, mae_scores, color='salmon', label='MAE')

# Ajout de titres et de légendes
plt.xlabel('Scores')
plt.title('Comparaison des performances des méthodes de prédiction')
plt.legend()

# Affichage du graphique
plt.show()

import matplotlib.pyplot as plt
import numpy as np

# Récupération des méthodes et des scores
methods = list(metrics.keys())
r2_scores = [metric['R2'] for metric in metrics.values()]
mse_scores = [metric.get('MSE', 0) for metric in metrics.values()]
mae_scores = [metric.get('MAE', 0) for metric in metrics.values()]

# Création des positions des barres
bar_width = 0.25
index = np.arange(len(methods))
opacity = 0.8

# Création de la figure et des sous-graphiques
fig, ax = plt.subplots(figsize=(12, 8))

# Barres pour R²
rects1 = ax.bar(index, r2_scores, bar_width, alpha=opacity, color='b', label='R²')

# Barres pour MSE
if any(mse_scores):
    rects2 = ax.bar(index + bar_width, mse_scores, bar_width, alpha=opacity, color='g', label='MSE')

# Barres pour MAE
if any(mae_scores):
    rects3 = ax.bar(index + 2 * bar_width, mae_scores, bar_width, alpha=opacity, color='r', label='MAE')

# Ajout des étiquettes, titres et légendes
ax.set_xlabel('Méthodes')
ax.set_ylabel('Scores')
ax.set_title('Comparaison des performances des méthodes de prédiction')
ax.set_xticks(index + bar_width)
ax.set_xticklabels(methods)
ax.legend()

# Affichage du graphique
plt.tight_layout()
plt.show()
