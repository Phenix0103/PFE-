import openpyxl

# Charger le fichier Excel
fichier_excel_path = 'C:/Users/ASUS TUF I5/Desktop/PFE/ETUDE DU SECTEUR/EtudeSB - Copie.xlsx'
fichier_excel = openpyxl.load_workbook(fichier_excel_path)

# Initialisation d'un dictionnaire pour stocker les totaux de prêts par année
totals_prets_numeriques_par_annee = {}

# Parcourir les années de 2010 à 2023 (pour inclure 2022)
for annee in range(2010, 2023):
    feuille_bilan = f'BILAN {annee}'
    # Vérifier si la feuille de bilan pour l'année donnée existe dans le fichier
    if feuille_bilan in fichier_excel.sheetnames:
        ws = fichier_excel[feuille_bilan]
        
        # Parcourir les lignes pour trouver "Créances sur clientèle"
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Créances sur clientèle" in row[0]:
                # Trouver la ligne des "Créances sur clientèle"
                # Supposer que le total sectoriel est la dernière valeur numérique de la ligne
                # Calculer le total des prêts pour l'année en sommant les valeurs numériques de la ligne,
                # en ignorant les cellules vides et la formule à la fin
                total_numerique = sum(val for val in row[1:] if isinstance(val, (int, float)))
                # Stocker le total dans le dictionnaire avec l'année comme clé
                totals_prets_numeriques_par_annee[annee] = total_numerique
                break  # Sortir de la boucle une fois la ligne trouvée

# Affichage des résultats
for annee, total in totals_prets_numeriques_par_annee.items():
    print(f"Total des prêts pour {annee}: {total} milliers de dinars")


import pandas as pd
import matplotlib.pyplot as plt

# Créer un DataFrame à partir des totaux des prêts
df_prets = pd.DataFrame(list(totals_prets_numeriques_par_annee.items()), columns=['Année', 'Total des Prêts (en milliers de dinars)'])

# Affichage du DataFrame
print(df_prets)

# Générer un graphique linéaire pour visualiser l'évolution des prêts de 2010 à 2022
plt.figure(figsize=(10, 6))
plt.plot(df_prets['Année'], df_prets['Total des Prêts (en milliers de dinars)'], marker='o', linestyle='-', color='blue')
plt.title('Évolution des prêts dans le secteur bancaire en Tunisie (2010-2022)')
plt.xlabel('Année')
plt.ylabel('Total des Prêts (en milliers de dinars)')
plt.grid(True)
plt.xticks(rotation=45)
plt.tight_layout()

# Sauvegarder le graphique
#graphique_path = "C:/Users/ASUS TUF I5/Desktop/PFE/ETUDE DU SECTEUR/evolution_prets_2010_2022.png"
#plt.savefig(graphique_path)

plt.show() #graphique_path

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error
import numpy as np

# Charger le fichier Excel (Ajustez le chemin selon votre environnement)
fichier_excel_path = 'C:/Users/ASUS TUF I5/Desktop/PFE/ETUDE DU SECTEUR/EtudeSB - Copie.xlsx'
fichier_excel = openpyxl.load_workbook(fichier_excel_path)

totals_prets_numeriques_par_annee = {}

for annee in range(2010, 2023):
    feuille_bilan = f'BILAN {annee}'
    if feuille_bilan in fichier_excel.sheetnames:
        ws = fichier_excel[feuille_bilan]
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Créances sur clientèle" in row[0]:
                total_numerique = sum(val for val in row[1:] if isinstance(val, (int, float)))
                totals_prets_numeriques_par_annee[annee] = total_numerique
                break

# Création d'un DataFrame
df_prets = pd.DataFrame(list(totals_prets_numeriques_par_annee.items()), columns=['Année', 'Total des Prêts'])

# Modélisation: Régression Linéaire
X = df_prets[['Année']]
y = df_prets['Total des Prêts']

# Diviser les données en ensemble d'apprentissage et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

# Entraîner le modèle
modele = LinearRegression()
modele.fit(X_train, y_train)

# Prédiction sur l'ensemble de test
predictions = modele.predict(X_test)

# Evaluation du modèle
mae = mean_absolute_error(y_test, predictions)
mse = mean_squared_error(y_test, predictions)
rmse = np.sqrt(mse)

print(f"MAE: {mae}, MSE: {mse}, RMSE: {rmse}")

# Visualisation des prêts réels vs prédits
plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='black', label='Données réelles')
plt.plot(X_test, predictions, color='blue', linewidth=2, label='Prédiction linéaire')
plt.title('Réel vs Prédit')
plt.xlabel('Année')
plt.ylabel('Total des Prêts')
plt.legend()
plt.show()

from sklearn.model_selection import train_test_split

# Préparation des variables indépendantes (X) et dépendantes (y)
X = df_prets[['Année']].values # Variable indépendante
y = df_prets['Total des Prêts'].values # Variable dépendante

# Séparation des données en ensembles d'entraînement et de test
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

from sklearn.linear_model import LinearRegression

# Création et entraînement du modèle
modele = LinearRegression()
modele.fit(X_train, y_train)

from sklearn.metrics import mean_squared_error, r2_score

# Prédiction sur les données de test
y_pred = modele.predict(X_test)

# Calcul et affichage des métriques d'évaluation
rmse = mean_squared_error(y_test, y_pred, squared=False)
r2 = r2_score(y_test, y_pred)

print(f"RMSE: {rmse}")
print(f"R² Score: {r2}")

plt.figure(figsize=(10, 6))
plt.scatter(X_test, y_test, color='black', label='Données réelles')
plt.plot(X_test, y_pred, color='blue', linewidth=3, label='Prédiction linéaire')
plt.title('Prédiction vs Réelles')
plt.xlabel('Année')
plt.ylabel('Total des Prêts (en milliers de dinars)')
plt.legend()
plt.show()
